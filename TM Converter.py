#System variable and io handling
import sys, getopt
import os
import configparser
import string
from libs.aiotranslator import Translator
#Object copy, handle excel style
import copy
import multiprocessing
from multiprocessing import Process, Queue, Pool
import queue 
import subprocess
#Get timestamp
import time
#function difination
import pyperclip
#GUI

#from tkinter import Tk
#from tkinter.ttk import *
from tkinter import *
from tkinter import Frame
from tkinter import filedialog
from tkinter import messagebox

version = '1.2d'
DELAY1 = 20



#**********************************************************************************
#	UI ******************************************************************
#**********************************************************************************

class SimpleTranslator(Frame):
	def __init__(self, Root = None, Result = None):
		
		Frame.__init__(self, Root) 
		self.parent = Root 
		self.Result = Result
		self.TMSource = None
		self.TMTarget = None
		self.DictionarySource = None
		self.DictionaryTarget = None
		self.initUI()

	def initUI(self):
	
		self.parent.title("Simple Translator " + version)
		self.to_language = ""
		self.from_language = ""
		
		Row = 1
		Label(self.parent, text='Convert to Workbook').grid(row=Row, column=1, columnspan=3, padx=10, pady=10, sticky=W)

		#Row = 2
		#Label(self.parent, textvariable= 'Translation Memory convert to Workbook').grid(row=Row, column=1, columnspan=5, padx=10, pady=10, sticky=E)
		#First row
		
		Row +=1
		Button(self.parent, text="Browse", width = 10, command= self.BrowseTMFile).grid(row=Row, column=1, padx=10, pady=10)
		self.CurrentTMFile = StringVar()
		self.TextFilePath = Entry(self.parent,text="Select your document", width=80, state="readonly", textvariable=self.CurrentTMFile)
		self.TextFilePath.grid(row=Row, column=2, padx=0, pady=10, sticky=W)
		Button(self.parent, text="Save", width = 10, command= self.SaveToWorkbook).grid(row=Row, column=3, padx=10, pady=10)
		

		#Row +=1
		#Label(self.parent, textvariable = 'Workbook convert to Translation Memory').grid(row=Row, column=1, columnspan=5, padx=10, pady=10, sticky=E)
		Row +=1
		Label(self.parent, text='Convert to Translation Memory').grid(row=Row, column=1, columnspan=3, padx=10, pady=10, sticky=W)

		#2nd row
		Row +=1
		Button(self.parent, text="Browse", width = 10, command= self.BrowseWorkbookFile).grid(row=Row, column=1, padx=10, pady=10)
		self.CurrentDictionaryFile = StringVar()
		self.TextFilePath = Entry(self.parent,text="Select your document", width=80, state="readonly", textvariable=self.CurrentDictionaryFile)
		self.TextFilePath.grid(row=Row, column=2, padx=0, pady=10, sticky=W)
		Button(self.parent, text="Save", width = 10, command= self.SaveToPickle).grid(row=Row, column=3, padx=10, pady=10)

	def CorrectPath(self, path):
		return str(path).replace('/', '\\')

	def BrowseTMFile(self):
		filename = filedialog.askopenfilename(title = "Select TM file",filetypes = (("Translation Memory","*.pkl"), ), )	
		if filename != "":
			self.TMSource = self.CorrectPath(filename)
			self.CurrentTMFile.set(str(self.TMSource))
		else:
			self.Notice.set("No document is selected")
			return

	def BrowseWorkbookFile(self):
		filename = filedialog.askopenfilename(title = "Select Workbook file",filetypes = (("Workbook","*.xlsx"), ), )	
		if filename != "":
			self.DictionarySource = self.CorrectPath(filename)
			self.CurrentTMFile.set(str(self.DictionarySource))
		else:
			self.Notice.set("No document is selected")
			return

	def SaveToWorkbook(self):
		if self.TMSource == None:
			self.Notice.set("Please select TM file before starting")
		elif not os.path.isfile(self.TMSource):
			self.Notice.set("TM file has been removed")
		else:	
			filename = filedialog.asksaveasfilename(title = "Save file to", filetypes = (("Worksheet", "*.xlsx"),),)
			if filename == "":
				self.Notice.set("Please enter a file name.")
			else:
				WorkbookFile = self.CorrectPath(filename)

				import pickle
				from openpyxl import load_workbook, worksheet, Workbook
				
				with open(self.TMSource, 'rb') as pickle_load:
					TM = pickle.load(pickle_load)
				if isinstance(TM, list):
					print('Old TM format')
					for Pair in TM:
						KO.append(Pair[0].lower())
						EN.append(Pair[1].lower())
					#self.en_tm = np.array(en)
					#self.ko_tm = np.array(ko)
				elif isinstance(TM, dict):
					print('New TM format')
					KO = TM['KO']
					EN = TM['EN']

				TempWB = Workbook()
				ws = TempWB.active
				ws.title = 'TranslationMemory'
				ws['A1'] = 'KO'
				ws['B1'] = 'EN'
				row = 2
				for i in range(len(EN)):
					ws['A'+ str(row)] = KO[i]
					ws['B'+ str(row)] = EN[i]
					row +=1
					
				TempWB.save(filename)
				print('Done')
	
	def SaveToPickle(self):
		if self.DictionarySource == None:
			self.Notice.set("Please select Workbook file before starting")
		elif not os.path.isfile(self.DictionarySource):
			self.Notice.set("Workbook file has been removed")
		else:	
			filename = filedialog.asksaveasfilename(title = "Save file to", filetypes = (("Translation Memory", "*.pkl"),),)
			if filename == "":
				self.Notice.set("Please enter a file name.")
			else:
				TMPath = self.CorrectPath(filename)

				import pickle
				from openpyxl import load_workbook, worksheet, Workbook
				
				xlsx = load_workbook(self.DictionarySource, data_only=True)

				EN = []
				KO = []
				TranslationMemory = {}
				for sheet in xlsx:
					EN_Coll = ""
					KR_Coll = ""
					database = None
					ws = xlsx[sheet.title]
					for row in ws.iter_rows():
						for cell in row:
							if cell.value == "KO":
								KR_Coll = cell.column_letter
								KR_Row = cell.row
								database = ws
							elif cell.value == "EN":
								EN_Coll = cell.column_letter
							if KR_Coll != "" and EN_Coll != "":
								print('Loading dictionary from: ', sheet.title)
								break	
						if database!=  None:
							break
					if database != None:
						for i in range(KR_Row, database.max_row): 
							KRAddress = KR_Coll + str(i+1)
							ENAddress = EN_Coll + str(i+1)
							KRCell = database[KRAddress]
							KRValue = KRCell.value
							ENCell = database[ENAddress]
							ENValue = ENCell.value
							if KRValue == None or ENValue == None or KRValue == 'KO' or ENValue == 'EN':
								continue
							elif KRValue != None and ENValue != None:
								EN.append(ENValue)
								KO.append(KRValue)
				TranslationMemory['EN'] = EN
				TranslationMemory['KO'] = KO
				with open(TMPath, 'wb') as pickle_file:
					pickle.dump(TranslationMemory, pickle_file, protocol=pickle.HIGHEST_PROTOCOL)
				print('Done')	

	def GenerateTranslatorEngine(self):
		self.Notice.set("Generate MyTranslator...")
		if self.Language.get() == 1:
			to_language = 'ko'
			from_language = 'en'
		else:
			to_language = 'en'
			from_language = 'ko'

		if self.TranslatorAgent.get() == 1:
			TranslatorAgent = 'google'
		else:
			TranslatorAgent = 'kakao'

		self.p1 = Process(target=GenerateTranslator, args=(self.MyTranslator_Queue, TranslatorAgent, from_language, to_language, self.DictionaryPath, self.ExceptionPath, self.TMPath,))
		self.p1.start()
		self.after(DELAY1, self.GetMyTranslator)
		return

	def GetMyTranslator(self):
		try:
			self.MyTranslator = self.MyTranslator_Queue.get_nowait()
		except queue.Empty:
			self.after(DELAY1, self.GetMyTranslator)
		if self.MyTranslator != None:
			self.Notice.set("Translator agent is created")
			self.TranslateBtn.configure(state=NORMAL)
			self.p1.join()
		else:
			self.Notice.set("Creating translator agent")
		return

# Function for processor	
def TranslationMemoryToDictionary(Mqueue, SourcePath = None, TargetPath = None):
	from openpyxl import load_workbook, worksheet, Workbook
	import pickle
	if SourcePath != None:
		if (os.path.isfile(SourcePath)):
			Start = time.time()
			with open(SourcePath, 'rb') as pickle_load:
				TM = pickle.load(pickle_load)
			End = time.time()
			Total = End - Start
			print('Total TM saved: ', len(TM))
			return TM
		else:
			return False
	else:
		return False
	
	#Mytranslator = Translator(TranslatorAgent, from_language, to_language, Dictionary, Exception, TranslationMemory)
	#Mqueue.put(Mytranslator)
	return

def main():
	Result = Queue()
	root = Tk()
	#root.geometry("400x350+300+300")
	SimpleTranslator(root, Result = Result)
	root.mainloop() 


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	dir_path = os.environ['APPDATA'] + '\\AOI Translator'
	appconfig = dir_path + '\\config.ini'
	BasePath = os.path.abspath(os.path.dirname(sys.argv[0]))
	Logger = BasePath +  "\\log.txt"
	#if (os.path.isfile(Logger)):


	if not os.path.isfile(appconfig):
		DefaultDatabase = BasePath +  "\\Database\\Dictionary.xlsx"
		if (os.path.isfile(DefaultDatabase)):
			Dictionary = DefaultDatabase
		else:
			Dictionary = None
		DefaultException = BasePath +  "\\Database\\Exception.xlsx"
		if (os.path.isfile(DefaultException)):
			Exception = DefaultException
		else:
			Exception = None
		DefaultTranslationMemory = BasePath +  "\\Database\\TranslationMemory.pkl"
		if (os.path.isfile(DefaultTranslationMemory)):
			TranslationMemory = DefaultTranslationMemory
		else:
			TranslationMemory = None	
	else:	
		config = configparser.ConfigParser()
		try:
			config.read(appconfig)
			try:
				Dictionary = config['Database']['path']
			except:
				Dictionary = BasePath +  "\\Database\\Exception.xlsx"
			try:
				Exception = config['Exception']['path']
			except:
				Exception = BasePath +  "\\Database\\Database.xlsx"
			try:
				TranslationMemory = config['TranslationMemory']['path']
			except:
				TranslationMemory = BasePath +  "\\Database\\TranslationMemory.pkl"

		except:
			pass
	main()
