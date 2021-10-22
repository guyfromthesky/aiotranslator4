import os, sys
from typing import Pattern
from openpyxl import load_workbook, worksheet, Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

basePath = os.path.abspath(os.path.dirname(sys.argv[0]))
FileName= input("Please enter your new file name: ")  #ask user to input the .xlsx file name
SheetName = input("Please enter your new sheet name: ") #ask user to inputh the .xlsx sheet name
NumSheet = input("Please enter how many sheets you want to create: ") #ask user to input the number of sheet they want to create.

BgFill = PatternFill(start_color='82F76F', end_color='82F76F', fill_type='solid')

wb = Workbook() #create excel file with 1 sheet.
wb.active.title = SheetName + "1"

#wb.remove(wb.active) #delete the default sheet 

j=int(NumSheet)+1 #var j to control number of created sheet in loop 

#start for loop to create the sheets 
for x in range(2,j):
	wb.create_sheet(SheetName+str(x)) #create new sheet

print(x)
ws =  wb.active

#start for loop to 
for y in range(0,int(NumSheet)):
	wb.active=y #Set new active sheet. 
	ws= wb.active
	print(wb.active)
	ws.cell(row=y+1,column=y+1).value = "This is sheet" + SheetName + str(y+1)
	
	ws.cell(row=y+1,column=y+1).fill = BgFill


try:
	wb.save(basePath + '//'+ FileName + '.xlsx') #try to save the new .xlsx file as FileName input.

except Exception as e:
	print('Failed to save the result: ' + str(e))

if (os.path.isfile(basePath + '//'+ FileName + '.xlsx')):
	print('File created successfully')
