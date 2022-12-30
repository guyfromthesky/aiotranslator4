#System variable and io handling
from genericpath import isfile
from msilib.schema import Extension
import sys
import os

#Regular expression handling
import re
# Multi-process support
from multiprocessing import Process , Queue, Manager, freeze_support
import queue 

from urllib.parse import urlparse

import subprocess
#Get timestamp
import time
from datetime import datetime
#function difination
import base64
import pickle
#import unicodedata

#from urllib.parse import urlparse

from ttkbootstrap.constants import *

#GUI
from ttkbootstrap import Entry, Label, Treeview, Scrollbar, OptionMenu
from ttkbootstrap import Checkbutton, Button, Notebook
from ttkbootstrap import Progressbar, Style, Window

from tkinter import Tk, Frame, Menubutton
from tkinter import Menu, filedialog, messagebox
from tkinter import Text
from tkinter import IntVar, StringVar
from tkinter import W, E, S, N, END, RIGHT, HORIZONTAL, NO, CENTER
from tkinter import WORD, NORMAL, BOTTOM, X, TOP, BOTH, Y, RAISED
from tkinter import DISABLED

from tkinter import scrolledtext 
from tkinter import simpledialog

from tkinter.dnd import *
# Web redirect
import webbrowser

from libs.aiotranslator import generate_translator

from libs.aioconfigmanager import ConfigLoader
from libs.documentprocessing import translate_docx, translate_msg
from libs.documentprocessing import translate_presentation, translate_workbook

from libs.general import get_version, resource_path, send_fail_request, get_user_name
from libs.tkinter_extension import AutocompleteCombobox, Generate_Document_Translate_Setting_UI, Apply_Transparency, Apply_FontSize

from google.cloud import logging
import pandas as pd

import pkgutil


tool_display_name = "Document Translator"
#This variable will be passed to AIO translator to know the source of translate request.
tool_name = 'document'
rev = 4202
ver_num = get_version(rev) 
version = tool_display_name  + " v" +  ver_num

DELAY = 20

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class DocumentTranslator(Frame):
	def __init__(self, Root, process_queue = None, result_queue = None, status_queue = None,
	my_translator_queue = None, my_db_queue = None, tm_manager = None, ):
		
		Frame.__init__(self, Root) 
		self.pack(side=TOP, expand=Y, fill=X)
		self.parent = Root 
		self.parent.protocol("WM_DELETE_WINDOW", self.on_closing)
		self.style =  Style()
		# Queue
		self.ProcessQueue = process_queue
		self.ResultQueue = result_queue
		self.StatusQueue = status_queue
		self.MyTranslator_Queue = my_translator_queue
		self.MyDB_Queue = my_db_queue
		# Shared memory between UI process and Translator proccess.
		# Unlike Queue.Queue(), Queue.Manager() support sharing object type. 
		self.TMManager = tm_manager
		# Translator Class
		self.MyTranslator = None
		

		self.Usage = 0
		self.Options = {}
		# Temporary Data
		self.ListFile = []
		# Tool option
		self.SourcePatch = ""
		self.LastDocument = ""
		#self.ShallowMode = ""
		self.SourceLanguage = ""
		self.Translator = ""
		#self.TurboMode = ""
		self.to_language = ""
		self.from_language = ""
		#self.Shallow = True	
		self.AppLanguage = "en"
		
		self.language_id_list = ['', 'ko', 'en', 'cn', 'jp', 'vi']
		self.language_list = ['', 'Hangul', 'English', 'Chinese', 'Japanese', 'Vietnamese']
		
		self.GlossaryList = []

		self.BUTTON_WIDTH = 20
		self.HALF_BUTTON_WIDTH = 15

		self.BUTTON_SIZE = 20
		self.HALF_BUTTON_SIZE = 15

		self.Btn_Style = "Accent.TButton"

		self.STATUS_LENGTH = 120

		self.LanguagePack = {}
		self.init_App_Setting()

		if self.AppLanguage != 'kr':
			from libs.languagepack import LanguagePackEN as LanguagePack
		else:
			from libs.languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack
	
		# Init function
		self.create_buttom_panel()

		self.init_theme()

		self.init_ui()
		
		self.init_UI_setting()

		#Create Translator
		if self.LicensePath.get() != "":
			if os.path.isfile(self.LicensePath.get()):
				self.generate_translator_engine()
			else:
				
				closed_box = messagebox.askokcancel('Bug Writer', 'No license selected, please select the key in Translate setting.',icon = 'info')

				if closed_box == True:
					self.TAB_CONTROL.select(self.TranslateSetting)
		else:
		
			closed_box = messagebox.askokcancel('Bug Writer', 'No license selected, please select the key in Translate setting.',icon = 'info')

			if closed_box == True:
				self.TAB_CONTROL.select(self.TranslateSetting)

		self.after(DELAY, self.debug_listening)	

	def create_buttom_panel(self):
		self.bottom_panel = BottomPanel(self)

	# UI init
	def init_ui(self):

		self.parent.resizable(False, False)
		self.parent.title(version)
		
		# Creating Menubar 
		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Info = StringVar()
		self.Progress = StringVar()
		
		self.Generate_DocumentTranslator_UI(self.MainTab)
		Generate_Document_Translate_Setting_UI(self, self.TranslateSetting)
		self.Generate_TM_Manager_UI(self.TM_Manager)
		self.Generate_DB_Uploader_UI(self.DB_Uploader)
		#self.Generate_Debugger_UI(self.Process)

		font_size  = self.Configuration['Document_Translator']['font_size']
		Apply_FontSize(font_size, self)

	def Generate_DocumentTranslator_UI(self, Tab):
		Row=1
		Label(Tab, textvariable=self.Progress, width= 40).grid(row=Row, column=1, columnspan=3, padx=5, pady=5, sticky=W)
		Label(Tab, textvariable=self.Notice, justify=RIGHT).grid(row=Row, column=3, columnspan=6, padx=5, pady=5, sticky= E)

		Row+=1
		Label(Tab, text=  self.LanguagePack.Label['Source'], width= 10, font='calibri 11 bold').grid(row=Row, column=1, padx=5, pady=5, sticky=W)

		self.source_language = StringVar()

		self.source_language_select = OptionMenu(Tab, self.source_language, *self.language_list, command = self.set_source_language)
		self.source_language_select.config(width=self.HALF_BUTTON_WIDTH)
		self.source_language_select.grid(row=Row, column=2, padx=0, pady=5, sticky=W)



		Label(Tab, text=  self.LanguagePack.Label['Language'], width= 10, font='calibri 11 bold').grid(row=Row, column=3, padx=5, pady=5, sticky=W)

		self.target_language = StringVar()

		self.target_language_select = OptionMenu(Tab, self.target_language, *self.language_list, command = self.set_target_language)
		self.target_language_select.config(width=self.HALF_BUTTON_WIDTH)
		self.target_language_select.grid(row=Row, column=4, padx=0, pady=5, sticky=W)

		'''
		self.Language = IntVar()	
		Radiobutton(Tab, width= 10, text=  self.LanguagePack.Option['Hangul'], value=1, variable=self.Language, command= self.set_target_language).grid(row=Row, column=2, padx=0, pady=5, sticky=W)
		Radiobutton(Tab, width= 10, text=  self.LanguagePack.Option['English'], value=2, variable=self.Language, command= self.set_target_language).grid(row=Row, column=3, padx=0, pady=5, sticky=W)
		'''
		Button(Tab, width = self.BUTTON_WIDTH, text=  self.LanguagePack.Button['Swap'], command= self.swap_language).grid(row=Row, column=7, padx=5, pady=5, sticky=E)		

		#Button(Tab, width = self.BUTTON_WIDTH, text=  self.LanguagePack.Button['RenewDatabase'], command= self.renew_my_translator).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.BUTTON_WIDTH, text=  self.LanguagePack.Button['OpenOutput'], command= self.OpenOutput).grid(row=Row, column=8, padx=5, pady=5, sticky=E)

		Row+=1
		Label(Tab, width= 10, text=  self.LanguagePack.Label['Source'], font='calibri 11 bold').grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.CurrentSourceFile = StringVar()
		self.TextFilePath = Entry(Tab,width = 128, text= self.LanguagePack.ToolTips['SelectSource'], state="readonly", textvariable=self.CurrentSourceFile)
		self.TextFilePath.grid(row=Row, column=2, columnspan=6, padx=5, pady=5, sticky=W+E)
		Button(Tab, width = self.BUTTON_WIDTH, text=  self.LanguagePack.Button['Browse'], command= self.BtnLoadDocument).grid(row=Row, column=8, columnspan=1, padx=5, pady=5, sticky=E)
		
		Row += 1

		# Menubutton Tool Option
		Extension_Option = Menubutton(Tab, text = self.LanguagePack.Label['ToolOptions'], width= 20, relief = RAISED)
		Extension_Option.grid(row=Row, column=2, padx=0, pady=5, sticky=W)
		Extension_Option_Menu = Menu(Extension_Option, tearoff = 0)

		self.TranslateFileName = IntVar()
		self.TranslateFileName.set(1)
		Extension_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['TranslateFileName'], variable = self.TranslateFileName)
		self.TranslateSheetName = IntVar()
		self.TranslateSheetName.set(1)
		Extension_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['TranslateSheetName'], variable = self.TranslateSheetName)

		self.SheetRemoval = IntVar() 
		self.SheetRemoval.set(0)
		Extension_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['SheetRemoval'], variable = self.SheetRemoval)

		self.DataOnly = IntVar()
		self.DataOnly.set(0)
		Extension_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['DataOnly'], variable = self.DataOnly)

		self.Bilingual = IntVar()
		self.Bilingual.set(0)
		Extension_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['Bilingual'], variable = self.Bilingual)

		Extension_Option["menu"] = Extension_Option_Menu

		# Menubutton TM Option
		TM_Option = Menubutton(Tab, text = self.LanguagePack.Label['TMOptions'], width= 20, relief = RAISED)
		TM_Option.grid(row=Row, column=3, padx=0, pady=5, sticky=W)
		TM_Option_Menu = Menu(TM_Option, tearoff = 0)

		self.TMUpdate = IntVar()
		self.TMUpdate.set(1)
		TM_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['UpdateTMFile'], variable = self.TMUpdate)

		self.TMTranslate = IntVar()
		self.TMTranslate.set(1)
		TM_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['TMTranslate'], variable = self.TMTranslate, command=self.TMTranslateModeToggle)
		
		TM_Option["menu"] = TM_Option_Menu


		# Menubutton ETC
		Result_File_Option = Menubutton(Tab, text = "Result Option:", width= 20, relief = RAISED)
		Result_File_Option.grid(row=Row, column=4, padx=0, pady=5, sticky=W)
		Result_File_Option_Menu = Menu(Result_File_Option, tearoff = 0)


		self.FixCorruptFileName = IntVar()
		self.FixCorruptFileName.set(1)	
		Result_File_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['CorruptFileName'], variable = self.FixCorruptFileName)

		#self.TurboTranslate = IntVar()
		#self.TurboTranslate.set(0)
		#Result_File_Option_Menu.add_checkbutton(label = self.LanguagePack.Option['TurboTranslate'], variable = self.TurboTranslate)

		Result_File_Option["menu"] = Result_File_Option_Menu

		Button(Tab, width = 20, text=  self.LanguagePack.Button['Stop'], command= self.Stop).grid(row=Row, column=7, columnspan=1,padx=5, pady=0, sticky=E)	
		self.btn_translate = Button(Tab, width = 20, text=  self.LanguagePack.Button['Translate'], command= self.Translate, state=DISABLED)
		self.btn_translate.grid(row=Row, column=8, columnspan=1, padx=5, pady=0, sticky=E)



		Row+=1

		Label(Tab, text="Sheet: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.SheetList = Text(Tab, width = 110, height=1) #
		self.SheetList.grid(row=Row, column=2, columnspan=7, padx=5, pady=5, sticky=E)

		Row+=1
		self.progressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.progressbar["maximum"] = 1000
		self.progressbar.grid(row=Row, column=1, columnspan=8, padx=5, pady=5, sticky=W)
		Row+=1
		self.Debugger = scrolledtext.ScrolledText(Tab, width=122, height=6, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, padx=5, columnspan=10, pady=5, sticky = E+W)

	def Generate_TranslateSetting_UI(self, Tab):
		Row = 1
		Label(Tab, text= self.LanguagePack.Label['LicensePath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TextLicensePath = Entry(Tab,width = 120, state="readonly", textvariable=self.LicensePath)
		self.TextLicensePath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.HALF_BUTTON_WIDTH, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_License_Path).grid(row=Row, column=8, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['TM']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TextLicensePath = Entry(Tab,width = 120, state="readonly", textvariable=self.TMPath)
		self.TextLicensePath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.HALF_BUTTON_WIDTH, text=  self.LanguagePack.Button['Browse'], command= self.SelectTM).grid(row=Row, column=8, columnspan=2, padx=5, pady=5, sticky=E)
		

	def Generate_Debugger_UI(self, Tab):	

		# Add clear console button.
		self.Debugger = scrolledtext.ScrolledText(Tab, width=122, height=17, undo=True, wrap=WORD, )
		self.Debugger.grid(row=0, column=0, padx=5, pady=5, sticky = E+W)


	### TM Manager tab ###
	def Generate_TM_Manager_UI(self, Tab):
		self.pair_list = []
		self.removed_list = []
		Max_Size = 10
		
		#Label(Tab, width = 100, text= "").grid(row=Row, column=1, columnspan=Max_Size, padx=5, pady=5, sticky = (N,S,W,E))
		Row = 1
		self.search_text = Text(Tab, height=1, width=100) #
		self.search_text.grid(row=Row,  column=1, columnspan=7, padx=5, pady=5, stick=W+E)

		#self.search_text.bind("<Enter>", self.search_tm_event)

		#print('Btn size', self.HALF_BUTTON_WIDTH)
		Button(Tab, text= self.LanguagePack.Button['Load'], width= self.HALF_BUTTON_WIDTH, command= self.load_tm_list).grid(row=Row,padx=5, pady=5, column=8, sticky=E)
		Button(Tab, text= self.LanguagePack.Button['Save'], width= self.HALF_BUTTON_WIDTH, command= self.save_tm).grid(row=Row,padx=5, pady=5, column=9, sticky=E)
		Button(Tab, width = self.HALF_BUTTON_WIDTH, text=  self.LanguagePack.Button['Search'] , command= self.search_tm_list).grid(row=Row,padx=5, pady=5, column=10,sticky=E)
		Row +=1
		#self.Debugger = Text(Tab, width=120, height=20, undo=True, wrap=WORD)
		#self.List = scrolledtext.ScrolledText(Tab, width=125, height=20, undo=True, wrap=WORD, )
		#self.List.grid(row=Row, column=1, columnspan=5, padx=5, pady=5)
		#style = Style()
		#style.configure('Treeview', background ="silver", foreground = "black")
		#style.map('Treeview', background = [('seleted', 'green')])
		
		self.Treeview = Treeview(Tab)
		self.Treeview.grid(row=Row, column=1, columnspan=11, padx=5, pady=5, sticky = N+S+W+E)
		verscrlbar = Scrollbar(Tab, orient ="vertical", command = self.Treeview.yview)
		#verscrlbar.pack(side ='right', fill ='x') 
		self.Treeview.configure(  yscrollcommand=verscrlbar.set)
	
		self.Treeview.Scrollable = True
		self.Treeview['columns'] = ('index', 'Source', 'Target')

		self.Treeview.column('#0', width=0, stretch=NO)
		self.Treeview.column('index', anchor=CENTER, width=0, stretch=NO)
		self.Treeview.column('Source', anchor=CENTER, width=130)
		self.Treeview.column('Target', anchor=CENTER, width=130)

		#self.source_tm_label = StringVar()
		#self.target_tm_label = StringVar()

		self.Treeview.heading('#0', text='', anchor=CENTER)
		self.Treeview.heading('index', text='index', anchor=CENTER)
		self.Treeview.heading('Source', text='Source', anchor=CENTER)
		self.Treeview.heading('Target', text='Target', anchor=CENTER)

		verscrlbar.grid(row=Row, column=11,  sticky = N+S+E)
		Tab.grid_columnconfigure(12, weight=0, pad=0)
		styles = Style()
		styles.configure('Treeview',rowheight=24)

		self.Treeview.bind("<Delete>", self.delete_treeview_line)	
		self.Treeview.bind("<Double-1>", self.double_right_click_treeview)	

		
		#Row +=1
		#self.Debugger = scrolledtext.ScrolledText(Tab, width=125, height=6, undo=True, wrap=WORD, )
		#self.Debugger.grid(row=Row, column=1, columnspan=Max_Size, padx=5, pady=5, sticky = (N,S,W,E))
 
	def Generate_DB_Uploader_UI(self, Tab):
		
		Row =1
		self.Str_DB_Path = StringVar()
		#self.Str_DB_Path.set('C:\\Users\\evan\\OneDrive - NEXON COMPANY\\[Demostration] V4 Gacha test\\DB\\db.xlsx')
		Label(Tab, text=  self.LanguagePack.Label['MainDB']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 130, state="readonly", textvariable=self.Str_DB_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=2, columnspan=6, padx=5, pady=5, sticky=W+E)
		Button(Tab, width = self.HALF_BUTTON_WIDTH, text=  self.LanguagePack.Button['Browse'], command= self.Btn_DB_Uploader_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['ProjectKey']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		
		self.ProjectList = AutocompleteCombobox(Tab)
		self.ProjectList.Set_Entry_Width(30)
		self.ProjectList.set_completion_list([])
		if self.glossary_id != None:
			self.ProjectList.set(self.glossary_id)

		self.ProjectList.grid(row=Row, column=2, columnspan=2, padx=5, pady=5, stick=W)
		Button(Tab, width = self.HALF_BUTTON_WIDTH, text= self.LanguagePack.Button['Reset'], command= self.Btn_DB_Uploader_Reset).grid(row=Row, column=7, columnspan=2,padx=5, pady=5, sticky=E)

		Button(Tab, width = self.HALF_BUTTON_WIDTH, text=  self.LanguagePack.Button['Update'], command= self.Btn_DB_Uploader_Execute_Script).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=E)

		Row += 1
		self.Uploader_Debugger = scrolledtext.ScrolledText(Tab, width=122, height=14, undo=True, wrap=WORD, )
		self.Uploader_Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)

	def debug_listening(self):
		while True:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Debugger.insert("end", "\r\n")
					ct = datetime.now()
					self.Debugger.insert("end", str(ct) + ": " + Status)
					self.Debugger.yview(END)
			except queue.Empty:
				break
		self.after(DELAY, self.debug_listening)

	def write_debug(self, text):
		try:
			self.Debugger.insert("end", "\n\r")
			ct = datetime.now()
			self.Debugger.insert("end", str(ct) + ": " + text)
			self.Debugger.yview(END)
		except:
			pass	

	def search_tm_event(self, event):
		self.search_tm_list()

	def delete_treeview_line(self, event):
		selected = self.Treeview.selection()
		to_remove = []
		for child_obj in selected:
			child = self.Treeview.item(child_obj)
			tm_index = child['values'][0]
			to_remove.append(tm_index)
			self.Treeview.delete(child_obj)
			
		#print('Current TM pair: ', len(self.MyTranslator.translation_memory))
		print('Current Dataframe pair: ', len(self.MyTranslator.current_tm))
		try:
			self.MyTranslator.current_tm = self.MyTranslator.current_tm.drop(to_remove)
		except Exception as e:
			print('Error:', e)
		
		print('After removed TM pair: ', len(self.MyTranslator.current_tm))
		#self.save_app_config()

	def double_right_click_treeview(self, event):
		focused = self.Treeview.focus()
		child = self.Treeview.item(focused)
		self.w
		self.Debugger.insert("end", "\r\n")
		self.Debugger.insert("end", 'Korean: ' + str(child["text"]))
		self.Debugger.insert("end", "\r\n")
		self.Debugger.insert("end", 'English: ' + str(child["values"][0]))
		self.Debugger.yview(END)
		#self.pair_list.delete("1.0", END)
		#self.pair_list.insert("end", text)
		#print(child)

	# Nam will check
	def load_tm_list(self):
		"""
		When clicking the [Load] button in TM Manager tab
		Display the pair languages in the text box.
		"""
		self.remove_treeview()
		tm_size = len(self.MyTranslator.translation_memory)
		
		self.Treeview.heading('Source', text='Source' + ' (' + self.MyTranslator.from_language.upper() + ') ', anchor=CENTER)
		self.Treeview.heading('Target', text='Target' + ' (' + self.MyTranslator.to_language.upper() + ') ',  anchor=CENTER)
		
		for index, pair in self.MyTranslator.translation_memory.iterrows():	
			from_str = pair[self.MyTranslator.from_language]
			to_str = pair[self.MyTranslator.to_language]
			if from_str != None:
				#print("Pair:", ko_str, en_str)
				try:
					#self.Treeview.insert('', 'end', text= str(pair['ko']), values=([str(pair['en'])]))
					self.Treeview.insert('', 'end', text= '', values=( index, str(from_str), str(to_str)))
					#print('Inserted id:', id)
				except:
					pass	
					

	def search_tm_list(self):
		"""
		Search text box in TM Manager tab
		Display the pair result from the text entered in the search field.
		"""
		text = self.search_text.get("1.0", END).replace("\n", "").replace(" ", "")
		self.remove_treeview()
		print("Text to search:", text)
		text = text.lower()
		if text != None:
			try:
				if len(self.MyTranslator.translation_memory) > 0:
					#translated = self.translation_memory[self.to_language].where(self.translation_memory[self.from_language] == source_text)[0]
					#result_from = self.MyTranslator.translation_memory[self.MyTranslator.translation_memory[self.MyTranslator.from_language].str.match(text)]
					#result_to = self.MyTranslator.translation_memory[self.MyTranslator.translation_memory[self.MyTranslator.to_language].str.match(text)]
					result_from = self.MyTranslator.translation_memory[self.MyTranslator.translation_memory[self.MyTranslator.from_language].str.contains(text)]
					result_to = self.MyTranslator.translation_memory[self.MyTranslator.translation_memory[self.MyTranslator.to_language].str.contains(text)]
					result = result_from.append(result_to)
					#print('type', type(result), 'total', len(result))
					if len(result) > 0:
						for index, pair in result.iterrows():
							#self.Treeview.insert('', 'end', text= str(pair['ko']), values=([str(pair['en'])]))
							self.Treeview.insert('', 'end', text= '', values=(index, str(pair[self.MyTranslator.to_language]), str(pair[self.MyTranslator.from_language])))
			except Exception  as e:
				#print('Error message (TM):', e)
				pass

	def remove_treeview(self):
		for i in self.Treeview.get_children():
			self.Treeview.delete(i)
	

	def save_tm(self):
		print('Saving config')
		UpdateProcess = Process(target=self.MyTranslator.export_current_translation_memory,)
		UpdateProcess.start()
		self.removed_list = []

	def Generate_Menu_UI(self):

		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		file = Menu(menubar, tearoff = 0)
		# Adding Load Menu 
		menubar.add_cascade(label =  self.LanguagePack.Menu['File'], menu = file) 
		file.add_command(label =  self.LanguagePack.Menu['SaveSetting'], command = self.save_app_config) 
		#file.add_command(label =  self.LanguagePack.Menu['LoadException'], command = self.SelectException) 
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['LoadTM'], command = self.SelectTM) 
		file.add_command(label =  self.LanguagePack.Menu['CreateTM'], command = self.SaveNewTM)
		file.add_separator() 
		file.add_command(label =  'Restart', command = self.rebuild_UI) 
		file.add_command(label =  self.LanguagePack.Menu['Exit'], command = self.on_closing)
		
		

		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.OpenWeb) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 

	def Generate_Tab_UI(self):

		MainPanel = Frame(self, name='mainpanel')
		MainPanel.pack(side=TOP, fill=BOTH, expand=Y)
		self.TAB_CONTROL = Notebook(MainPanel, name='notebook')
		# extend bindings to top level window allowing
		#   CTRL+TAB - cycles thru tabs
		#   SHIFT+CTRL+TAB - previous tab
		#   ALT+K - select tab using mnemonic (K = underlined letter)
		self.TAB_CONTROL.enable_traversal()

		#self.TAB_CONTROL = Notebook(self.parent)
		#Tab1
		self.MainTab = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.MainTab, text= self.LanguagePack.Tab['Main'])
		#Tab2
		self.TranslateSetting = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.TranslateSetting, text= self.LanguagePack.Tab['Translator'])
		#Tab5
		#self.Comparison = ttk.Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.Comparison, text=  self.LanguagePack.Tab['Comparison'])
		#Tab6
		self.TM_Manager = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.TM_Manager, text= self.LanguagePack.Tab['TMManager'])

		self.DB_Uploader = Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.DB_Uploader, text= self.LanguagePack.Tab['DBUploader'])

		# Process Details tab
		#self.Process = Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.Process, text = self.LanguagePack.Tab['Debug'])

		#self.TAB_CONTROL.pack(expand=1, fill="both")
		self.TAB_CONTROL.pack(side=TOP, fill=BOTH, expand=Y)

	def disable_button(self):
		_state = DISABLED
		self.bottom_panel.btn_renew_translator.configure(state=_state)
		self.bottom_panel.project_id_select.configure(state=_state)
		self.btn_translate.configure(state=_state)
		
	def enable_button(self):
		_state = NORMAL
		self.bottom_panel.btn_renew_translator.configure(state=_state)
		self.bottom_panel.project_id_select.configure(state=_state)
		self.btn_translate.configure(state=_state)

	# Menu Function
	def About(self):
		messagebox.showinfo("About....", "Creator: Giang - evan@nexonnetworks.com")

	def Error(self, ErrorText):
		messagebox.showerror('Tool error...', ErrorText)	

	def SaveAppLanguage(self, language):
		print('Save language:', language)
		self.write_debug(self.LanguagePack.ToolTips['AppLanuageUpdate'] + " "+ language) 
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'app_lang', language)

	def SaveAppTransparency(self, transparency):
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'Transparent', transparency)

	def save_app_config(self):
		target_language = self.target_language.get()
		target_language_index = self.language_list.index(target_language)
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'target_lang', target_language_index)
		
		source_language = self.source_language.get()
		source_language_index = self.language_list.index(source_language)
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'source_language', source_language_index)

		#self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'speed_mode', self.TurboTranslate.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'bilingual', self.Bilingual.get())

		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'value_only', self.DataOnly.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'file_name_correct', self.TranslateFileName.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'file_name_translate', self.TranslateFileName.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'sheet_name_translate', self.TranslateSheetName.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'tm_translate', self.TMTranslate.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'tm_update', self.TMUpdate.get())
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'remove_unselected_sheet', self.SheetRemoval.get())

	def swap_language(self):
		
		source_language = self.source_language.get()
		source_language_index = self.language_list.index(source_language)
		target_language = self.target_language.get()
		target_language_index = self.language_list.index(target_language)
		self.target_language.set(source_language)
		self.source_language.set(target_language)

		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'target_lang', source_language_index)
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'source_lang', target_language_index)
		self.MyTranslator.set_language_pair(target_language = self.language_id_list[source_language_index], source_language = self.language_id_list[target_language_index])
		#self._dictionary_status.set(str(len(self.MyTranslator.dictionary)))
		self.TMStatus.set(str(self.MyTranslator.translation_memory_size))

	def SetLanguageKorean(self):
		self.AppLanguage = '1'
		self.SaveAppLanguage(self.AppLanguage)
		self.rebuild_UI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = '2'
		self.SaveAppLanguage(self.AppLanguage)
		self.rebuild_UI()

	def OpenWeb(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/%5BTranslation%5D+AIO+Translator")

	def on_closing(self):
		if messagebox.askokcancel("Quit", "Do you want to quit?"):
			self.parent.destroy()
			try:
				
				self.TranslatorProcess.terminate()
			except Exception as e:
				print(e)

	def rebuild_UI(self):
		if messagebox.askokcancel("Quit", "Do you want to restart?"):
			self.parent.destroy()
			main()
		else:
			messagebox.showinfo('Language update','The application\'s language will be changed in next session.')	

	def CorrectPath(self, path):
		if sys.platform.startswith('win'):
			return str(path).replace('/', '\\')
		else:
			return str(path).replace('\\', '//')
	
	def init_theme(self):
		"""Applied the theme name saved in the settings on init."""
		print('init_theme')
		try:
			all_themes = self.style.theme_names()
			personalize_themes = ['wynnmeister', 'erza\'s', 'tien\'s', 'dao\'s', 'blackpink']
			self.theme_names = []
			for theme in all_themes:
				
				if theme in personalize_themes:
					print('Personalize theme:', theme)
					user = get_user_name()
					if theme == 'dao\'s' and user == 'jennie':
						self.theme_names.append(theme)
					elif theme == 'wynnmeister' and user == 'wynn.saltywaffle':
						self.theme_names.append(theme)	
					elif theme == 'erza\'s' and user == 'erzaerza':
						self.theme_names.append(theme)	
					elif theme == 'tien\'s' and user == 'hann':
						self.theme_names.append(theme)	
					elif theme == 'blackpink' and user == 'ruko':
						self.theme_names.append(theme)
					elif user == 'evan':
						self.theme_names.append(theme)
				else:	
					self.theme_names.append(theme)
	

		

							#	['cosmo', 'flatly', 'litera', 'minty',
							#	"lumen", "sandstone",	"yeti", "pulse", 
							#	"united", "morph", "journal", "darkly", 'superhero', 
							#	'solar', 'cyborg', 'vapor', 'simplex', 'cerculean'
							#		, 'pinky']
			if self.used_theme not in self.theme_names:
				raise Exception('Cannot use the theme saved in the config'
					' because it is not supported or required files have'
					' been removed.')

			self.style.theme_use(self.used_theme)

		except Exception as err:
			print('Error while initializing theme:\n'
				f'- {err}\n'
				'The system default theme will be used instead.')
		transparency  = self.Configuration['Document_Translator']['Transparent']
		Apply_Transparency(transparency, self)

		


	def select_theme_name(self):
		"""Save the theme name value to Configuration and change
		the theme based on the selection in the UI.
		
		Args:
			config_theme_name -- str
				Theme name retrieved from config. (Default: '')
		"""
		try:
			theme_name = self.strvar_theme_name.get()
			print('Select theme:', theme_name)
			self.style.theme_use(theme_name)
			self.AppConfig.Save_Config(
				self.AppConfig.Doc_Config_Path,
				'Document_Translator',
				'theme_name',
				theme_name, True)

		except Exception as err:
			messagebox.showerror(
				title='Error',
				message=f'Error occurs when selecting theme: {err}')

	def remove_theme(self):
		print('remove_theme')
		"""Remove the theme saved in config then restart the app."""
		self.AppConfig.Save_Config(
			self.AppConfig.Doc_Config_Path,
			'Document_Translator',
			'theme_name',
			'')
		
		messagebox.showinfo(
			title='Info',
			message='App will restart to apply the change.')
		self.parent.destroy()
		main()

	def CorrectExt(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename = os.path.splitext(baseName)[0]
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath


	def Btn_Select_License_Path(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectDB'],filetypes = (("JSON files","*.json" ), ), )	
		if filename != "":
			LicensePath = self.CorrectPath(filename)
			self.AppConfig.Save_Config(self.AppConfig.Translator_Config_Path, 'Translator', 'license_file', LicensePath, True)
			os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = LicensePath
			self.LicensePath.set(LicensePath)
			self.rebuild_UI()
		else:
			self.write_debug("No file is selected")


	def SelectTM(self):
		filename = filedialog.askopenfilename(title = "Select Translation Memory file", filetypes = (("TM files","*.pkl"), ),)
		if filename != "":
			NewTM = self.CorrectPath(filename)
			self.TMPath.set(NewTM)
			self.AppConfig.Save_Config(self.AppConfig.Translator_Config_Path, 'Translator', 'translation_memory', NewTM, True)
			self.renew_my_translator()
			self.write_debug(self.LanguagePack.ToolTips['TMUpdated'])
		else:
			self.write_debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
			

	def SaveNewTM(self):	
		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Translation Memory", "*.pkl"),),)
		filename = self.CorrectExt(filename, "pkl")
		if filename == "":
			return
		else:
			NewTM = self.CorrectPath(filename)
			with open(NewTM, 'wb') as pickle_file:
				# New TM format.
				pickle.dump({'tm_version': 4}, pickle_file, protocol=pickle.HIGHEST_PROTOCOL)
			self.TMPath.set(NewTM)
			self.AppConfig.Save_Config(self.AppConfig.Translator_Config_Path, 'Translator', 'translation_memory', NewTM, True)
			self.renew_my_translator()

	def _save_project_key(self, event):
		glossary_id = self.bottom_panel.project_id_select.get()
		glossary_id = glossary_id.replace('\n', '')

		self.AppConfig.Save_Config(self.AppConfig.Translator_Config_Path, 'Translator', 'glossary_id', glossary_id)
		self.MyTranslator.glossary_id = glossary_id
		
		self.renew_my_translator()

	def OpenOutput(self):
		if len(self.ListFile) > 0:
			Source = self.ListFile[0]
			Outputdir = os.path.dirname(Source)
			BasePath = str(os.path.abspath(Outputdir))
			os.startfile(BasePath)
		else:
			self.Error('No file selected')	
		
	def BtnLoadDocument(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("All type files","*.docx *.xlsx *.xlsm *.pptx *.msg"), ("Workbook files","*.xlsx *.xlsm"), ("Document files","*.docx"), ("Presentation files","*.pptx"), ("Email files","*.msg"), ("PDF files","*.pdf")), multiple = True)	
		if filename != "":
			self.ListFile = list(filename)
			self.CurrentSourceFile.set(str(self.ListFile[0]))
			self.write_debug(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.write_debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def BtnLoadRawSource(self):
		#filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files","*.xlsx *.xlsm *xlsb"), ("Document files","*.docx")), multiple = True)	
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.RawFile = FolderName
			self.RawSource.set(str(FolderName))
			
			self.write_debug(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.write_debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return


###########################################################################################
# DB UPLOADER 
###########################################################################################

	def Btn_DB_Uploader_Browse_DB_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)	
		if filename != "":
			self.DB_Path = self.CorrectPath(filename)
			self.Str_DB_Path.set(self.DB_Path)
			self.write_debug(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.write_debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_DB_Uploader_Reset(self):
		glossary_id = self.ProjectList.get()
		result = self.Confirm_Popup(glossary_id, 'Please type \''+ glossary_id + "\' to confirm.")
		
		if result == True:
			self.Generate_DB_Reset_Processor = Process(target=function_reset_db, args=(self.StatusQueue, self.ResultQueue, self.MyTranslator, glossary_id))
			self.Generate_DB_Reset_Processor.start()
			self.after(DELAY, self.Wait_For_Reset_Processor)	

		
	def Wait_For_Reset_Processor(self):
		"""
		After [Execute] button processing
		Display process information in the text box.
		"""
		if (self.Generate_DB_Reset_Processor.is_alive()):
			self.after(DELAY, self.Wait_For_Reset_Processor)
		else:
			try:
				self.Uploader_Debugger.insert("end", "\n\r")
				self.Uploader_Debugger.insert("end", "CSV DB is reset")
			except queue.Empty:
				pass
			self.Generate_DB_Reset_Processor.terminate()


	def Btn_DB_Uploader_Execute_Script(self):
		glossary_id = self.ProjectList.get()
		result = self.Confirm_Popup(glossary_id, 'Please type \''+ glossary_id + "\' to confirm.")
		
		if result == True:
			DB = self.Str_DB_Path.get()
			self.Generate_DB_Processor = Process(target=function_create_csv_db, args=(self.StatusQueue, self.ResultQueue, DB))
			self.Generate_DB_Processor.start()
			self.after(DELAY, self.Wait_For_Creator_Processor)	
	
	def Wait_For_Creator_Processor(self):
		"""
		After [Execute] button processing
		Display process information in the text box.
		"""
		db_path = None
		if (self.Generate_DB_Processor.is_alive()):

			try:
				db_path = self.ResultQueue.get(0)
				if db_path != False:
		
					self.Uploader_Debugger.insert("end", "\n\r")
					self.Uploader_Debugger.insert("end", "CSV DB is generated")
					self.Uploader_Debugger.insert("end", "\n\r")
					self.Uploader_Debugger.insert("end", "Compare generated DB with the current version")
			except queue.Empty:
				pass	
			self.after(DELAY, self.Wait_For_Creator_Processor)
		else:
			try:
				db_path = self.ResultQueue.get(0)
				if db_path != False:
					self.Uploader_Debugger.insert("end", "\n\r")
					self.Uploader_Debugger.insert("end", "CSV DB is generated")
					self.Uploader_Debugger.insert("end", "\n\r")
					self.Uploader_Debugger.insert("end", "Compare generated DB with the current version")
			except queue.Empty:
				pass
			self.Generate_DB_Processor.terminate()
		if db_path != None:	
			glossary_id = self.ProjectList.get()
			self.Compare_DB_Processor = Process(target=function_compare_db, args=(self.StatusQueue, self.ResultQueue, self.MyTranslator, glossary_id, db_path))
			self.Compare_DB_Processor.start()
			self.after(DELAY, self.Wait_For_DB_Compare_Processor)	
		'''
		MsgBox = messagebox.askquestion ('Exit Application','Are you sure you want to exit the application',icon = 'warning')
		if MsgBox == 'yes':
			self.Generate_DB_Processor = Process(target=function_create_csv_db, args=(self.ResultQueue, self.MyTranslator))
			self.AuGenerate_DB_Processortomation_Processor.start()
			self.after(DELAY, self.Wait_For_Uploader_Processor)	
		else:
			messagebox.showinfo('Return','You will now return to the application screen')
		'''
		
	def Wait_For_DB_Compare_Processor(self):
		"""
		After wait_for_creator_processor() process
		Display the old and new DB comparison in the text box.
		"""
		compare_result = None
		if (self.Compare_DB_Processor.is_alive()):
			self.after(DELAY, self.Wait_For_DB_Compare_Processor)
		else:
			try:
				compare_result = self.ResultQueue.get(0)
				self.Uploader_Debugger.insert("end", "\n\r")
				self.Uploader_Debugger.insert("end", "Wait for user's confirmation.")
			except queue.Empty:
				pass
			self.Compare_DB_Processor.terminate()
		_new_added = True
		if compare_result != None:
			if compare_result == False:

				if self.Supper != True:	
					message = 'You are not permisted to upload new project. If you feel this is a mistake, please contact your leader.'
					self.Uploader_Debugger.insert("end", "\n\r")
					self.Uploader_Debugger.insert("end", message)
					messagebox.showerror('Error', message)
					return 
				else:
					message = 'Are you sure you want to upload this new project?'

			else:
				message = 'Are you sure you want to upload the DB?'
				_new_added = False
				
			message += '\n'
			change_flag = False
			if _new_added != True:
				if compare_result['dropped']>0:
					message += 'Dropped: ' + str(compare_result['dropped']) + '\n'
					change_flag = True
				if compare_result['added']>0:
					message += 'Added: ' + str(compare_result['added']) + '\n'
					change_flag = True
				if compare_result['changed']>0:
					message += 'Changes: ' + str(compare_result['changed']) + '\n'
					change_flag = True
				if change_flag == False:
					message += 'OLD and NEW DB are identical.'
				
			self.Uploader_Debugger.insert("end", "\n\r")
			self.Uploader_Debugger.insert("end", message)
			
			result = messagebox.askquestion ('Notice',message,icon = 'warning')
			if result == 'yes':
				glossary_id = self.ProjectList.get()
				self.Upload_DB_Processor = Process(target=function_upload_db, args=(self.StatusQueue, self.ResultQueue,self.MyTranslator, glossary_id, compare_result))
				self.Upload_DB_Processor.start()
				self.after(DELAY, self.Wait_For_Uploader_Processor)	
			else:
				self.Uploader_Debugger.insert("end", "\n\r")
				self.Uploader_Debugger.insert("end", 'Canceled, no change is made.')
		
	def Wait_For_Uploader_Processor(self):
		"""
		After wait_for_db_compare_processor() process
		Display information whether the upload is successful in the text box.
		"""
		if (self.Upload_DB_Processor.is_alive()):
			self.after(DELAY, self.Wait_For_Uploader_Processor)
		else:
			self.Uploader_Debugger.insert("end", "\n\r")
			try:
				upload_result = self.ResultQueue.get(0)	
				if upload_result == "False":
					self.Uploader_Debugger.insert("end", "Fail to upload DB")
				elif upload_result == 'Forbidden':
					self.Uploader_Debugger.insert("end", "License file hss no permission")
				elif upload_result == 'LostDB':
					self.Uploader_Debugger.insert("end", "WARNING: Project ID is removed from the list, please try to upload the DB again.")
				else:
					self.Uploader_Debugger.insert("end", "DB updated")
			except queue.Empty:
				self.Uploader_Debugger.insert("end", "DB updated")

			self.Upload_DB_Processor.terminate()
	
	def Confirm_Popup(self, Request, message):
		MsgBox = simpledialog.askstring(title="Confirmation popup", prompt=message)
		self.Supper = False
		secret = Request + '_' + "SU"
		if MsgBox == secret:
			self.Supper = True
			return True
		if MsgBox == Request:
			return True
		else:
			return False

###########################################################################################


	def onExit(self):
		self.quit()

	def init_App_Setting(self):

		self.LicensePath = StringVar()

		self.Transparent = IntVar()
		self.FontSize = IntVar()

		self.DictionaryPath = StringVar()
		self.TMPath = StringVar()

		self.CurrentDataSource = StringVar()
		self.Notice = StringVar()
		
		
		self.TMStatus  = StringVar()
		#self.HeaderStatus = StringVar()
		self._dictionary_status = StringVar()
		self._version_status  = StringVar()
		self._update_day = StringVar()

		self.AppConfig = ConfigLoader(Document = True)
		self.Configuration = self.AppConfig.Config
		_app_language = self.Configuration['Document_Translator']['app_lang']
		if _app_language == 1:
			self.AppLanguage  = 'kr'
		else:
			self.AppLanguage  = 'en'
		
		license_file_path = self.Configuration['Translator']['license_file']
		self.LicensePath.set(license_file_path)
		os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = license_file_path

		_tm_path = self.Configuration['Translator']['translation_memory']
		
		Transparent  = self.Configuration['Document_Translator']['Transparent']
		self.Transparent.set(Transparent)

		# Check if tm path is valid
		if not os.path.exists(_tm_path):
			with open(_tm_path, 'wb') as pickle_file:
				pickle.dump({'tm_version': 4}, pickle_file, protocol=pickle.HIGHEST_PROTOCOL)
				pickle_file.close()
		self.TMPath.set(_tm_path)		

		self.bucket_id = self.Configuration['Translator']['bucket_id']
		self.db_list_uri = self.Configuration['Translator']['db_list_uri']
		self.project_bucket_id = self.Configuration['Translator']['project_bucket_id']

		self.glossary_id = self.Configuration['Translator']['glossary_id']

		self.used_theme = self.Configuration['Document_Translator']['theme_name']
		self.font_size = self.Configuration['Document_Translator']['font_size']		

		self.strvar_theme_name = StringVar()


	def init_UI_setting(self):

		print('Config:', self.Configuration)

		self.target_language.set(self.language_list[self.Configuration['Document_Translator']['target_lang']])
		self.source_language.set(self.language_list[self.Configuration['Document_Translator']['source_lang']])
		
		#self.TurboTranslate.set(self.Configuration['Document_Translator']['speed_mode'])
		self.Bilingual.set(self.Configuration['Document_Translator']['bilingual'])
		self.DataOnly.set(self.Configuration['Document_Translator']['value_only'])
		self.FixCorruptFileName.set(self.Configuration['Document_Translator']['file_name_correct'])
		self.TranslateFileName.set(self.Configuration['Document_Translator']['file_name_translate'])
		self.TranslateSheetName.set(self.Configuration['Document_Translator']['sheet_name_translate'])
		self.TMTranslate.set(self.Configuration['Document_Translator']['tm_translate'])
		self.TMUpdate.set(self.Configuration['Document_Translator']['tm_update'])
		self.SheetRemoval.set(self.Configuration['Document_Translator']['remove_unselected_sheet'])
		
		try:
			self.Usage = self.Configuration['Document_Translator']['usage']
		except:
			self.Usage = 0

	def renew_my_translator(self):
		self.MyTranslator = None
		self.disable_button()
		self.generate_translator_engine()


	def set_target_language(self, target_language):
		index = self.language_list.index(target_language)
		to_language = self.language_id_list[index]
	
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'target_lang', index)

		self.MyTranslator.set_target_language(to_language)
		#self._dictionary_status.set(str(len(self.MyTranslator.dictionary)))
		self.TMStatus.set(str(self.MyTranslator.translation_memory_size))
		

	def set_source_language(self, source_language):
		index = self.language_list.index(source_language)
		from_language = self.language_id_list[index]

		self.MyTranslator.from_language
		self.AppConfig.Save_Config(self.AppConfig.Doc_Config_Path, 'Document_Translator', 'source_lang', index)

		self.MyTranslator.set_source_language(from_language)	
		#self._dictionary_status.set(str(len(self.MyTranslator.dictionary)))
		self.TMStatus.set(str(self.MyTranslator.translation_memory_size))


	
	def Stop(self):
		try:
			if self.TranslatorProcess.is_alive():
				self.TranslatorProcess.terminate()
		except:
			pass
		self.progressbar["value"] = 0
		self.progressbar.update()
		self.write_debug('Translate Process has been stop')
		return

	def Stop(self):
		try:
			if self.TranslatorProcess.is_alive():
				self.TranslatorProcess.terminate()
		except:
			pass
		self.progressbar["value"] = 0
		self.progressbar.update()
		self.write_debug('Translate Process has been stop')	

	def generate_translator_engine(self):
		self.write_debug(self.LanguagePack.ToolTips['AppInit'])

		target_language = self.language_id_list[self.language_list.index(self.target_language.get())]
		source_language = self.language_id_list[self.language_list.index(self.source_language.get())]

		self.glossary_id = self.bottom_panel.project_id_select.get()
		self.glossary_id = self.glossary_id.replace('\n', '')
		tm_path = self.TMPath.get()
		print('Start new process: Generate Translator')
		self.TranslatorProcess = Process(	target = generate_translator,
											kwargs= {	'my_translator_queue' : self.MyTranslator_Queue, 
														'temporary_tm' : self.TMManager, 
														'from_language' : source_language, 
														'to_language' : target_language, 
														'glossary_id' : self.glossary_id, 
														'used_tool' : tool_name,
														'tm_path' : tm_path,
														'bucket_id' : self.bucket_id, 
														'db_list_uri' : self.db_list_uri, 
														'project_bucket_id' : self.project_bucket_id,
													},
										)
		self.TranslatorProcess.start()
		self.after(DELAY, self.GetMyTranslator)
		return

	def GetMyTranslator(self):
		try:
			st = time.time()
			self.MyTranslator = self.MyTranslator_Queue.get_nowait()
			#print('Get Translator',time.time()- st)
		except queue.Empty:
			self.after(DELAY, self.GetMyTranslator)

		#print("self.MyTranslator: ", self.MyTranslator)	
		if self.MyTranslator != None:	
			#print("My translator is created")
	
			self.enable_button()

			self._dictionary_status.set(str(self.MyTranslator.glossary_size))
			self.TMStatus.set(str(self.MyTranslator.translation_memory_size))

			glossary_list = [""] + self.MyTranslator.glossary_list
			bucket_db_list = [""] + self.MyTranslator.bucket_db_list
			self.bottom_panel.project_id_select.set_completion_list(glossary_list)
			self.ProjectList.set_completion_list(bucket_db_list)

			if self.glossary_id in self.MyTranslator.glossary_list:
				self.bottom_panel.project_id_select.set(self.glossary_id)
			else:
				self.bottom_panel.project_id_select.set("")

			if self.glossary_id in self.MyTranslator.bucket_db_list:
				self.ProjectList.set(self.glossary_id)
			else:
				self.ProjectList.set(self.glossary_id)
				#self.Error('No Valid Project selected, please update the project key and try again.')	
			
			if isinstance(self.MyTranslator.version, str):
				version = self.MyTranslator.version[0:10]
			else:
				version = '-'

			if isinstance(self.MyTranslator.update_day, str):
				Date = self.MyTranslator.update_day[0:10]
			else:
				Date = '-'


			self.TMPath.set(str(self.MyTranslator.tm_path))
			self._version_status.set(version)
			self._update_day.set(Date)
			
			self.write_debug(self.LanguagePack.ToolTips['AppInitDone'])
			all_tm = self.MyTranslator.all_tm
			self.write_debug('TM version: ' +  str(all_tm['tm_version']))
			self.write_debug('TM sub version: ' + str(all_tm['tm_sub_version']))
			#all_key = all_tm.keys()
			#print('All project key: ', all_tm.keys())
			for key in all_tm:
				if key == 'tm_version':
					pass
				elif key == 'tm_sub_version':
					pass
				else:
					self.write_debug(key + ": " + str(len(all_tm[key])))
			
			self.TranslatorProcess.join()
			
	

	def TMTranslateModeToggle(self):
		if self.TMTranslate.get() == 1:
			self.MyTranslator.tm_translate_enable(True)
		else:
			self.MyTranslator.tm_translate_enable(False)

	def GetOptions(self):
		#Get and set language
		target_language = self.language_id_list[self.language_list.index(self.target_language.get())]
		source_language = self.language_id_list[self.language_list.index(self.source_language.get())]
		
		if source_language == target_language:
			messagebox.showinfo('Error', 'Source and Target language is the same.')	
			return False
		self.MyTranslator.set_language_pair(target_language = target_language, source_language = source_language)
		
		#Add Subscription key
		#self.MyTranslator.SetSubscriptionKey(self.SubscriptionKey)	

		#Set TM Update Mode
		if self.TMUpdate.get() == 1:
			self.MyTranslator.tm_update_anable(True)
		else:
			self.MyTranslator.tm_update_anable(False)

		#Set Predict mode 
		#if self.TurboTranslate.get() == 1:
		#	self.MyTranslator.source_language_predict_enable(True)
		#else:
		#	self.MyTranslator.source_language_predict_enable(False)

		if self.Bilingual.get() == 1:
			self.Options['Bilingual']  = True
		else:
			self.Options['Bilingual'] = False

		#Set Data Mode
		if self.DataOnly.get() == 1:
			self.Options['DataOnly']  = True
		else:
			self.Options['DataOnly'] = False	

		#Set Skip Mode
		'''
		if self.SkipEmptyRow.get() == 1:
			self.Options['SkipMode'] = True
		else:
			self.Options['SkipMode'] = False
		'''
		self.Options['SkipMode'] = False
		#Set Sheet removal mode
		if self.SheetRemoval.get() == 1:
			self.Options['SheetRemovalMode'] = True
		else:
			self.Options['SheetRemovalMode'] = False

		# Set Update Mode
		if self.TMUpdate.get() == 1:
			self.Options['TMUpdateMode'] = True
		else:
			self.Options['TMUpdateMode'] = False



		#Set TM Translate Mode
		#print('self.TMTranslate', self.TMTranslate.get())
		if self.TMTranslate.get() == 1:
			self.Options['MemoryTranslate']  = True
		else:
			self.Options['MemoryTranslate'] = False

		#Set Translate File name option
		if self.TranslateFileName.get() == 1:
			self.Options['TranslateFileName'] = True
		else:
			self.Options['TranslateFileName'] = False
		
		#Set Translate Sheet name option
		if self.TranslateSheetName.get() == 1:
			self.Options['TranslateSheetName'] = True
		else:
			self.Options['TranslateSheetName'] = False

		#Set Name fixer mode
		if self.FixCorruptFileName.get() == 1:
			self.Options['FixCorruptFileName'] = True
		else:
			self.Options['FixCorruptFileName'] = False	

		#Get sheet list
		try:
			Raw = self.SheetList.get("1.0", END).replace("\n", "").replace(" ", "")	
			self.Options['Sheet'] = Raw.split(",")
		except:
			self.Options['Sheet'] = []

		# Get Documents list
		self.Options['SourceDocument'] = self.ListFile
		if self.Options['SourceDocument'] == "":
			self.write_debug(self.LanguagePack.ToolTips['SourceNotSelected']) 
			return
		else:
			self.write_debug(self.LanguagePack.ToolTips['DocumentLoad']) 


	def Translate(self):

		result = self.GetOptions()
		if result == False:
			return
		self.progressbar["value"] = 0
		self.progressbar.update()
		#SourceDocument = self.TextFilePath.get()
		
		
		
		# Kill exist Translator Process
		try:
			if self.TranslatorProcess.is_alive():
				self.TranslatorProcess.terminate()
		except:
			pass

		self.disable_button()

		# Clear exist queue
		try:
			while True:
				percent = self.ProcessQueue.get_nowait()
				print("Remain percent: ", percent)
		except queue.Empty:
			pass

		#self.TranslatorProcess = Process(target=Execuser, args=(self.MyTranslator, self.ProcessQueue ,self.ResultQueue, self.StatusQueue, SourceDocument, 
		#Multiple, TranslateFileName, TranslateSheetName, FixCorruptFileName, Sheet, DataMode, SheetRemovalMode, TMUpdateMode, SkipMode, TMTranslate,LastDocument, ))
		self.TranslatorProcess = Process(target=execute_document_translate, args=(self.MyTranslator, self.ProcessQueue ,self.ResultQueue, self.StatusQueue, self.Options,))
		self.TranslatorProcess.start()
		self.after(DELAY, self.GetCompleteStatus)

	def GetCompleteStatus(self):
		if (self.TranslatorProcess.is_alive()):
			try:
				percent = self.ProcessQueue.get(0)
				self.progressbar["value"] = percent
				self.progressbar.update()
				self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass
			
			self.after(DELAY, self.GetCompleteStatus)
		else:
			try:
				Result = self.ResultQueue.get(0)		
				if Result == True:
					self.progressbar["value"] = 1000
					self.progressbar.update()	
					self.Progress.set("Progress: " + str(100) + '%')
					
				elif '[Errno 13] Permission denied' in Result:
					self.write_debug(self.LanguagePack.ToolTips['TranslateFail'])
					Result = Result.replace('[Errno 13] Permission denied','File is being in used')
					self.Error(str(Result))
					self.progressbar["value"] = 0
					self.progressbar.update()
					self.Progress.set("Progress: " + str(0) + '%')
				elif 'Package not found at' in Result:
					self.write_debug(self.LanguagePack.ToolTips['TranslateFail'])
					Result = Result.replace(' Package not found at ','File is being in used: ')
					self.Error(str(Result))
					self.progressbar["value"] = 0
					self.progressbar.update()
					self.Progress.set("Progress: " + str(0) + '%')
				else:	
					self.write_debug(self.LanguagePack.ToolTips['TranslateFail'])
					self.Error(str(Result))
					self.progressbar["value"] = 0
					self.progressbar.update()
					self.Progress.set("Progress: " + str(0) + '%')

				self.TranslatorProcess.terminate()
				self.enable_button()

			
			except queue.Empty:
				self.enable_button()
			while True:
				try:
					percent = self.ProcessQueue.get(0)
					self.progressbar["value"] = percent
					self.progressbar.update()
					self.Progress.set("Progress: " + str(percent/10) + '%')
				except queue.Empty:
					break

			if self.TMUpdate.get() == 1:
				self.renew_my_translator()
				#self.TMStatus.set(str(self.MyTranslator.translation_memory_size))

class BottomPanel(Frame):
	def __init__(self, master):
		Frame.__init__(self, master) 
		self.pack(side=BOTTOM, fill=X)          # resize with parent
		
		# separator widget
		#Separator(orient=HORIZONTAL).grid(in_=self, row=0, column=1, sticky=E+W, pady=5)
		Row = 1
		Col = 1
		Label(text='Update', width=15).grid(in_=self, row=Row, column=Col, padx=5, pady=5, stick=E)
		Col += 1
		Label(textvariable=master._update_day, width=15).grid(in_=self, row=Row, column=Col, padx=0, pady=5, stick=E)
		master._update_day.set('-')
		Col += 1
		DictionaryLabelA = Label(text=master.LanguagePack.Label['Database'], width=15)
		DictionaryLabelA.grid(in_=self, row=Row, column=Col, padx=5, pady=5, stick=E)
		Col += 1
		Label(textvariable=master._dictionary_status, width=15).grid(in_=self, row=Row, column=Col, padx=0, pady=5, stick=E)
		master._dictionary_status.set('-')
		Col += 1
		TMLabel=Label(text=  master.LanguagePack.Label['TM'], width=15)
		TMLabel.grid(in_=self, row=Row, column=Col, padx=5, pady=5, sticky=E)
		TMLabel.bind("<Enter>", lambda event : master.write_debug(master.LanguagePack.ToolTips['FilePath'] + master.TMPath.get()))
		Col += 1
		Label( width= 15, textvariable=master.TMStatus).grid(in_=self, row=Row, column=Col, padx=0, pady=5, sticky=E)
		master.TMStatus.set('-')

		Col += 1
		Label(text='Project', width=15).grid(in_=self, row=Row, column=Col, padx=5, pady=5, stick=E)

		Col += 1
		self.project_id_select = AutocompleteCombobox()
		self.project_id_select.Set_Entry_Width(25)
		self.project_id_select.set_completion_list([])
		self.project_id_select.grid(in_=self, row=Row, column=Col, padx=5, pady=5, stick=E)
		if master.glossary_id != None:
			self.project_id_select.set(master.glossary_id )

		self.project_id_select.bind("<<ComboboxSelected>>", master._save_project_key)
		Col += 1
		self.btn_renew_translator = Button(text=master.LanguagePack.Button['RenewDatabase'], width=15, command= master.renew_my_translator, state=DISABLED)
		self.btn_renew_translator.grid(in_=self, row=Row, column=Col, padx=10, pady=5, stick=E)
		
		
		self.rowconfigure(0, weight=1)
		self.columnconfigure(0, weight=1)


def Optimize(SourceDocument, StatusQueue):
	from openpyxl import load_workbook, worksheet, Workbook
	from openpyxl.styles import Font
	#print(SourceDocument)
	FileList = os.listdir(SourceDocument)
	#print(FileList)
	for FileName in FileList:

		if FileName != None:
			File = SourceDocument + '//' + FileName
			try:
				xlsx = load_workbook(File, data_only=True)
			except:
				continue
			
			Outputdir = os.path.dirname(File) + '/Optimized/'
			baseName = os.path.basename(File)
			sourcename, ext = os.path.splitext(baseName)
			TranslatedName	= sourcename

			StatusQueue.put('Optimizing file: ' + TranslatedName)
			

			if not os.path.isdir(Outputdir):
				try:
					os.mkdir(Outputdir)
					output_file = Outputdir + sourcename + ext
				except OSError:
					print ("Creation of the directory %s failed" % Outputdir)
					output_file = os.path.dirname(File) + '/' + sourcename + '_Optmized' + ext
			else:
				output_file = Outputdir + sourcename + ext	
			
			try:
				xlsx.save(output_file)
				#StatusQueue.put('Optimized done.')
			except Exception as e:
				StatusQueue.put('Failed to save the result: ' + str(e))
	StatusQueue.put('Optimized done.')	


###########################################################################################

###########################################################################################
def function_reset_db(status_queue, result_queue, MyTranslator, glossary_id):
	print('Reset DB')
	result = MyTranslator.reset_glossary(glossary_id)
	print('function_reset_db', result)
	if result == False:
		status_queue.put('Fail to reset DB')
	else:
		status_queue.put("DB reset.")

def function_create_csv_db(status_queue, result_queue, db_path):
	try:
		result_path = function_create_db_data(db_path)
		result_queue.put(result_path)
	except Exception as e:
		status_queue.put('Error while creating csv db: ' + str(e))
	
def function_compare_db(status_queue, result_queue, MyTranslator, glossary_id, address):

	# The project may exist, but was not registered. Return false result
	if glossary_id not in MyTranslator.glossary_list:
		result = {
			'dropped': False,
			'added': False,
			'changed': False,
			'path': address
		}
		result_queue.put(result)
		return

	# Download the DB of the projects from the server
	try:
		new_csv_path = address['db']
		sourcename, ext = os.path.splitext(new_csv_path)
		old_csv_db = sourcename + '_old' +  ext
		try:
			result = MyTranslator.download_db_to_file(glossary_id, old_csv_db)
			if result == False:
				status_queue.put('Fail to load OLD DB file')
				status_queue.put(str(_error_message))
				result_queue.put(False)
				return
				
		except Exception as _error_message:
			status_queue.put('Fail to load OLD DB file')
			status_queue.put(str(_error_message))
			result_queue.put(False)
			return
		
		if not isfile(old_csv_db):
			result = {
				'dropped': 0,
				'added': 0,
				'changed': 0,
				'path': address
			}
			status_queue.put('OLD DB file is not existed.')
			result_queue.put(False)
			return


		# Read and compare the old and new DB.
		old_db = pd.read_csv(old_csv_db, usecols = ['en', 'ko', 'vi', 'ja', 'zh-TW'])
		new_db = pd.read_csv(new_csv_path, usecols = ['en', 'ko', 'vi', 'ja', 'zh-TW'])
		old_db['version'] = "old"
		new_db['version'] = "new"

		old_cols = old_db.columns.tolist()
		Old_Index_ID = old_cols[1]
		#new_cols = new_db.columns.tolist()

		old_accts_all = set(old_db[Old_Index_ID])
		new_accts_all = set(new_db[Old_Index_ID])

		dropped_accts = old_accts_all - new_accts_all
		added_accts = new_accts_all - old_accts_all

		all_data = pd.concat([old_db,new_db], ignore_index=True)
		changes = all_data.drop_duplicates(subset=None, keep= 'last')
		dupe_accts = changes[changes[Old_Index_ID].duplicated() == True][Old_Index_ID].tolist()
		dupes = changes[changes[Old_Index_ID].isin(dupe_accts)]
		
		change_new = dupes[(dupes["version"] == "new")]
		change_old = dupes[(dupes["version"] == "old")]
		change_new = change_new.drop(['version'], axis=1)
		change_old = change_old.drop(['version'], axis=1)

		change_new.set_index(Old_Index_ID, inplace=True)
		change_new = change_new.fillna("#NA")

		change_old.set_index(Old_Index_ID, inplace=True)
		change_old = change_old.fillna("#NA")
		
		# Combine all the changes together
		try:			
			df_all_changes = pd.concat([change_old, change_new],
										axis='columns',
										keys=['old', 'new'],
										join='outer')
			df_all_changes = df_all_changes.swaplevel(axis='columns')[change_new.columns[0:]]
			#df_all_changes.fillna("#NA")
			df_changed = df_all_changes.groupby(level=0, axis=1).apply(lambda frame: frame.apply(report_diff, axis=1))
			#create a list of text columns (int columns do not have '{} ---> {}')
			df_changed = df_changed.reset_index()
			
			df_changed['has_change'] = df_changed.apply(has_change, axis=1)
			#df_changed.tail()
			diff = df_changed[(df_changed.has_change == 'Y')]
			diff = diff.reindex(columns=Old_Index_ID)
		except:
			diff = []
		result = {
			'dropped': len(dropped_accts),
			'added': len(added_accts),
			'changed': len(diff),
			'path': address
		}

	except Exception as e:
		result = {
				'dropped': 0,
				'added': 0,
				'changed': 0,
				'path': address
			}
		status_queue.put("Error when compare DB: " + str(e))

	result_queue.put(result)

def has_change(row):
	if "-->" in row.to_string():
		return "Y"
	else:
		return "N"

def report_diff(x):
	#print(x)
	if len(x) == 2:
		if x[0] == x[1]:
			#return x[0]
			return x[0]
		elif x[0] == "#NA":
			return "[ADD] --> " + str(x[1])
		elif x[1]== "#NA":
			return "[DROP] --> " + str(x[0])
		else:
			return '{} --> {}'.format(*x)
	else:
		#return x[0]
		return x[0]

def function_upload_db(status_queue, result_queue, MyTranslator, glossary_id, result):
	print('Upload DB')
	print(result)

	address = result['path']
	add = result['added']
	drop = result['dropped']
	changes = result['changed']

	result = MyTranslator.backup_and_update_blob(glossary_id, address)
	print('backup_and_update_blob', result)
	if result == False:
		status_queue.put('Fail to upload DB')
		result_queue.put("False")
		return
	elif result == 'Forbidden':
		status_queue.put('No permission to upload DB')
		result_queue.put("Forbidden")
		return
	elif result == 'LostDB':
		status_queue.put('Project ID was Removed but couldn\'t be created.')
		result_queue.put("LostDB")
		return

	client = logging.Client()
	
	log_name = 'db-update'

	logger = client.logger(log_name)
	
	try:
		if sys.platform.startswith('win'):
			try:
				user_name = os.getlogin()
			except Exception as Error_msg:
				print('Error:', Error_msg)
				user_name = os.environ['COMPUTERNAME']
		else:
			try:
				user_name = os.environ['LOGNAME']
			except Exception as Error_msg:
				print('Error:', Error_msg)
				user_name = "Anonymous"
	except Exception as Error_msg:
		print('Error:', Error_msg)
		user_name = "Anonymous"

	data_object = {
		'tool': 'Document Translator',
		'translator_ver': ver_num,
		'glossary_id': glossary_id,
		'db_file': str(address['db']),
		'added': add,
		'dropped': drop,
		'changed': changes,
	}

	tracking_object = {
		'user': user_name,
		'details': data_object
	}
	
	try:
		logger.log_struct(tracking_object)
	except Exception  as e:
		print('exception:', e)
		result = False
	
	status_queue.put("DB updated.")

def get_datestamp():
	_now = datetime.now()	
	_date_time = _now.strftime("%m/%d/%Y")
	return _date_time


# Load the DB from xlsx file and return DB object:
# db_object['info'] = @dict
# db_object['db'] = @dict

def function_create_db_data(DB_Path):
	print('Create DB from:', DB_Path)
	from openpyxl import load_workbook
	import csv

	DatabasePath = DB_Path

	Outputdir = os.path.dirname(DatabasePath)
	baseName = os.path.basename(DatabasePath)
	sourcename, ext = os.path.splitext(baseName)

	#output_file = Outputdir + '/' + sourcename + '_SingleFile.xlsx'
	output_db_csv = Outputdir + '/' + sourcename + '_db'+ '.csv'
	output_header_csv = Outputdir + '/' + sourcename + '_header'+ '.csv'
	output_info_csv = Outputdir + '/' + sourcename + '_info'+ '.csv'
	SpecialSheets = ['info']

	if DatabasePath != None:
		
		if (os.path.isfile(DatabasePath)):
			print('Load DB from file: ', DatabasePath)
			xlsx = load_workbook(DatabasePath)
			DictList = []
			#Dict = []
			with open(output_db_csv, 'w', newline='', encoding='utf_8_sig') as csv_db, open(output_header_csv, 'w', newline='', encoding='utf_8_sig') as csv_header, open(output_info_csv, 'w', newline='', encoding='utf_8_sig') as csv_info:
				db_writer = csv.writer(csv_db, delimiter=',')
				db_writer.writerow(['','ko', 'en', 'zh-TW', 'ja', 'vi', 'description'])
				
				header_writer = csv.writer(csv_header, delimiter=',')
				header_writer.writerow(['ko', 'en', 'cn', 'jp', 'vi', 'description'])
				
				info_writer = csv.writer(csv_info, delimiter=',')
				info_writer.writerow(['date', get_datestamp()])
				
				print('Looking for DB in each sheet:')
				for sheet in xlsx:
					sheetname = sheet.title.lower()
					print('Current sheet:', sheetname)
					if sheetname not in SpecialSheets:
						# init loop
						list_col = {}
						list_col['EN'] = ""
						list_col['KO'] = ''
						list_col['CN'] = ''
						list_col['JP'] = ''
						list_col['VI'] = ''

						start_row = 0

						database = None
						ws = xlsx[sheet.title]

						for row in ws.iter_rows():
							language_count = 0
							for cell in row:
								cell_value = cell.value
								
								if cell_value in list_col:
									list_col[cell_value] = cell.column_letter
									start_row = cell.row
									language_count+=1
								
							if language_count > 1:
								database = ws
								DictList.append(sheet.title)
								break	
							
						if database != None:
							
							for i in range(start_row, database.max_row): 
								db_entry = {}
								valid = False
								for language in list_col:
									#print('Current language: ', language)
									if list_col[language] != '':
										
										cell_adress = list_col[language] + str(i+1)
										try:
											raw_cell_value = str(database[cell_adress].value)
										except:
											raw_cell_value = None	

										if raw_cell_value not in ['', None]:
											valid = True
											cell_value = raw_cell_value.replace('\r', '').replace('\n', '')	

											#if sheetname != 'header':
											#	cell_value = cell_value.lower()
											# Obsoleted
											# cell_value = basse64_encode(cell_value)
											#print('Encrypt value: ', cell_value)
										else:
											cell_value = ''	
										db_entry[language] = cell_value
									else:
										db_entry[language] = ''
								if valid:
									if sheetname != 'header':
										db_writer.writerow(['', db_entry['KO'], db_entry['EN'].lower(), db_entry['CN'], db_entry['JP'], db_entry['VI'].lower(), sheetname])
									elif sheetname != 'info':
										header_writer.writerow([db_entry['KO'], db_entry['EN'], db_entry['CN'], db_entry['JP'], db_entry['VI'], sheetname])
								#db_object['db'][sheetname].append(db_entry)

				
	_db = pd.read_csv(output_db_csv)
	_supported_language = []
	for language in ['en', 'ko', 'vi', 'ja', 'zh-TW']:
		if not _db[language].dropna().empty:
			#info_writer.writerow(['language', language])
			_supported_language.append(language)
	_db = _db[['en', 'ko', 'vi', 'ja', 'zh-TW']]
	_db = _db.drop_duplicates()
	_db.to_csv(output_db_csv, encoding ='utf_8_sig' )
				

	_address = {}
	_address['db'] = output_db_csv
	_address['header'] = output_header_csv
	_address['info'] = output_info_csv
	_address['language'] = _supported_language
	print('Create CSV DB completed.')
	return _address

def basse64_encode(string_to_encode):
	if string_to_encode == '':
		return string_to_encode
	raw_encoded_string =  str(base64.b64encode(string_to_encode.encode('utf-8')))
	encoded_string = re.findall(r'b\'(.+?)\'', raw_encoded_string)[0]
	
	return encoded_string

def base64_decode(string_to_decode):
	decoded_string = base64.b64decode(string_to_decode).decode('utf-8')
	
	return


###########################################################################################

def execute_document_translate(MyTranslator, ProgressQueue, ResultQueue, StatusQueue, Options,):
	print('Creating process for Translator...')
	Start = time.time()

	SourceDocument = Options['SourceDocument']


	Preflix = ""
	Preflix	+= MyTranslator.to_language.upper() + '_'
	
	Result = False

	if Options['TMUpdateMode']:
			#print(MyTranslator.LanguagePack.ToolTips['TMUpdating']) 
			#MyTranslator.ProactiveTMTranslate = False
			MyTranslator.tm_update_anable = True
	try:
		ProgressQueue.get_nowait()
	except queue.Empty:
		pass
		
	translate_file = []	

	for File in SourceDocument:
		if File != None:
			
			Outputdir = os.path.dirname(File)
			baseName = os.path.basename(File)
			sourcename, ext = os.path.splitext(baseName)

			try:
				Newsourcename = sourcename.encode('cp437').decode('euc_kr')
				StatusQueue.put('Correct malformed name: ' + Newsourcename)
				
			except:
				Newsourcename = sourcename
				#StatusQueue.put('fail to create new name: ' + Newsourcename)

			if Options['FixCorruptFileName'] == True:	
				try:
					fixed_name = sourcename.encode('cp437').decode('euc_kr')
					newPath = Outputdir + '/'+ fixed_name + ext
				except:
					newPath = File
				try:
					os.rename(File, newPath)
					StatusQueue.put('Fixed name: ' + fixed_name)
				except:
					newPath = File
			else:
				newPath = File

			now = datetime.now()

			timestamp = str(int(datetime.timestamp(now)))

			if Options['TranslateFileName']:
				Translated = MyTranslator.translate(Newsourcename)
				if Translated != False:
					TranslatedName = re.sub(r'[\\/\:*"<>\|\.%\$\^&£]', '', Translated)
				else:
					TranslatedName = Newsourcename
			else:
				TranslatedName	= Newsourcename
			if '.xls' not in ext:
				output_file = Outputdir + '/' + Preflix + TranslatedName + '_' + timestamp + ext
			else:
				output_file = Outputdir + '/' + Preflix + TranslatedName + '_' + timestamp + '.xlsx'
			#print('Output file name: ', output_file)
		
		Options['SourceDocument'] = newPath
		Options['OutputDocument'] = output_file
		print('Option:', Options)
		if ext == '.docx':
			Result = translate_docx(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			try:
				print('Debug')
				#Result = translate_docx(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			except Exception as e:
				ErrorMsg = ('Error message: ' + str(e))
				StatusQueue.put(str(ErrorMsg))
				Result = str(e)
				
		elif ext in ['.xlsx', '.xlsm']:
			#Result = translate_workbook(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			try:
				Result = translate_workbook(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			except Exception as e:
				ErrorMsg = ('Error message: ' + str(e))
				StatusQueue.put(str(ErrorMsg))
				Result = str(e)
		elif ext == '.msg':
			#Result = translate_msg(ProgressQueue=ProgressQueue, ResultQueue=ResultQueue, StatusQueue=StatusQueue, Mytranslator=MyTranslator, Options=Options)
			try:
				Result = translate_msg(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			except Exception as e:
				ErrorMsg = ('Error message: ' + str(e))

				StatusQueue.put(str(ErrorMsg))
				Result = str(e)

		elif ext == '.pptx':
			#Result = translate_presentation(ProgressQueue=ProgressQueue, ResultQueue=ResultQueue, StatusQueue=StatusQueue, Mytranslator=MyTranslator, Options=Options)
			try:
				Result = translate_presentation(progress_queue=ProgressQueue, result_queue=ResultQueue, status_queue=StatusQueue, MyTranslator=MyTranslator, Options=Options)
			except Exception as e:
				ErrorMsg = ('Error message: ' + str(e))
				StatusQueue.put(str(ErrorMsg))

				Result = str(e)

		if Result == True:
			translate_file.append(baseName)
			ResultQueue.put(Result)
			End = time.time()
			Total = End - Start
			StatusQueue.put('Total time spent: ' + str(Total) + ' second.')
			StatusQueue.put('Translated file: ' + Preflix + ' ' + TranslatedName + '_' + timestamp + ext)
		else:
			Message = 'Fail to translate document, details: \n' + str(Result)
			ResultQueue.put(str(Message))
	try:
		mem_tm = len(MyTranslator.temporary_tm)
		newTM = MyTranslator.append_translation_memory()
		MyTranslator.send_tracking_record(file_names = translate_file)
		StatusQueue.put('Source: ' + str(baseName))
		StatusQueue.put('TM usage: ' + str(MyTranslator.last_section_tm_request))
		StatusQueue.put('API usage: ' + str(MyTranslator.last_section_api_usage))
		StatusQueue.put('Invalid request: ' + str(MyTranslator.last_section_invalid_request))
		StatusQueue.put('TM In-memory: ' + str(mem_tm))
		StatusQueue.put('TM append this section: ' + str(newTM))
		
	except Exception as e:
		StatusQueue.put('Error while appending TM: ' + str(e))

def main():
	ProcessQueue = Queue()
	ResultQueue = Queue()
	StatusQueue = Queue()
	MyTranslatorQueue = Queue()
	
	MyManager = Manager()
	TMManager = MyManager.list()
	MyDB = Queue()
    
	root = Window(themename="minty")
	now = datetime.now()
	init_time = datetime.timestamp(now)	
	#application = DocumentTranslator(root, process_queue = ProcessQueue, result_queue = ResultQueue, status_queue = StatusQueue, my_translator_queue = MyTranslatorQueue, my_db_queue = MyDB, tm_manager = TMManager)
	try:
		application = DocumentTranslator(root, process_queue = ProcessQueue, result_queue = ResultQueue, status_queue = StatusQueue, my_translator_queue = MyTranslatorQueue, my_db_queue = MyDB, tm_manager = TMManager)
		
		icon_path = resource_path('resource/document_translator.ico')
		print(icon_path)
		if os.path.isfile(icon_path):
			root.iconbitmap(icon_path)
		now = datetime.now()
		complete_time = datetime.timestamp(now)	
		launch_time = (complete_time - init_time)  / 1000
		if launch_time > 120:
			MsgBox = messagebox.askquestion (title='Whitelist problem', message="Your application take long time to initial, please prefer to this link for the solution:\nhttps://confluence.nexon.com/display/NWVNQA/%5BTranslation+Tool%5D+Anti-virus+issue")
			if MsgBox == 'yes':
				webbrowser.open_new(r"https://confluence.nexon.com/display/NWVNQA/%5BTranslation+Tool%5D+Anti-virus+issue")
		else:
			print('Launch time:', launch_time)
		root.mainloop()	
	except Exception as e:
		try:
			root.withdraw()
		except:
			return
		try:
			send_fail_request(e, ver_num)
		except Exception as e2:
			messagebox.showinfo(title='Fail to launch the application', message="Error details has not been reported.\n Please contact with me (evan) if you need urgent support.")
			return
		#application.MyTranslator.write_local_log('Critical error: ' + str(e))
		messagebox.showinfo(title='Fail to launch the application', message="Error details has been reported.\n Please contact with me (evan) if you need urgent support.")

if __name__ == '__main__':
	if sys.platform.startswith('win'):
		freeze_support()
	main()
