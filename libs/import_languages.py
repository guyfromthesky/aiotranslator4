from pkgutil import get_data
from openpyxl import load_workbook
from io import BytesIO


'''
language_pack = pkgutil.get_data(_file_name, 'language.xlsx' )
xlsx = BytesIO(language_pack)
wb = load_workbook(xlsx)
'''
def print_workbook():
    wb = load_workbook('src/language.xlsx')
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value != None:
                    current_string = str(cell.value)
                    print('current_string', current_string)

