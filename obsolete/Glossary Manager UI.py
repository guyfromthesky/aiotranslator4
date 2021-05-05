#System variable and io handling
import sys, getopt
import os
import multiprocessing
from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp
import time
from datetime import datetime
import configparser

#GUI
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkinter import colorchooser
#from tkinter import style

from openpyxl import load_workbook, worksheet, Workbook

#MyTranslatorAgent = 'google'
Tool = "Automation Execuser"
VerNum = '0.0.1a'
version = Tool  + " " +  VerNum
DELAY1 = 20


class DataEntry:
	def __init__(self, String_EN = "", String_KR = "", Path = "", CWD = ""):
		#self.StringID = StringID
		self.String_EN = String_EN
		self.String_KR = String_KR
		self.Path = Path
		self.CWD = CWD

class MainDB:
	def __init__(self, SheetName, DB = {}):
		self.SheetName = SheetName
		self.DB = DB

class AutocompleteCombobox(Combobox):

	def set_completion_list(self, completion_list):
		"""Use our completion list as our drop down selection menu, arrows move through menu."""
		self._completion_list = sorted(completion_list, key=str.lower) # Work with a sorted list
		self._hits = []
		self._hit_index = 0
		self.position = 0
		self.bind('<KeyRelease>', self.handle_keyrelease)
		self['values'] = self._completion_list  # Setup our popup menu
		#self._w = 10
		self.delete(0,END)	

	def Set_Entry_Width(self, width):
		self.configure(width=width)

	def Set_DropDown_Width(self, width):
		print('Change size: ', width)
		style = Style()
		style.configure('TCombobox', postoffset=(0,0,width,0))
		self.configure(style='TCombobox')

	def autocomplete(self, delta=0):
		"""autocomplete the Combobox, delta may be 0/1/-1 to cycle through possible hits"""
		if delta: # need to delete selection otherwise we would fix the current position
			self.delete(self.position, END)
		else: # set position to end so selection starts where textentry ended
			self.position = len(self.get())
		# collect hits
		_hits = []
		for element in self._completion_list:
			if element.lower().startswith(self.get().lower()): # Match case insensitively
				_hits.append(element)
		# if we have a new hit list, keep this in mind
		if _hits != self._hits:
			self._hit_index = 0
			self._hits=_hits
		# only allow cycling if we are in a known hit list
		if _hits == self._hits and self._hits:
			self._hit_index = (self._hit_index + delta) % len(self._hits)
		# now finally perform the auto completion
		if self._hits:
			self.delete(0,END)
			self.insert(0,self._hits[self._hit_index])
			self.select_range(self.position,END)

	def handle_keyrelease(self, event):
		"""event handler for the keyrelease event on this widget"""
		if event.keysym == "BackSpace":
			self.delete(self.index(INSERT), END)
			self.position = self.index(END)
		if event.keysym == "Left":
			if self.position < self.index(END): # delete the selection
				self.delete(self.position, END)
			else:
				self.position = self.position-1 # delete one character
				self.delete(self.position, END)
		if event.keysym == "Right":
			self.position = self.index(END) # go to end (no selection)
		if len(event.keysym) == 1:
			self.autocomplete()

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class Glossary_Manager(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		#super().__init__()
		self.parent = Root 

		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']

		self.Manager = Manager['Default_Manager']

		self.Config_Init()

		self.Options = {}

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		self.DB_Path = ""
		self.Test_Case_Path = ""
		
		self.App_LanguagePack = {}
		self.initGeneralSetting()

		if self.AppLanguage != 'kr':
			from languagepack import LanguagePackEN as LanguagePack
		else:
			from languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack

		# Init function

		self.parent.resizable(False, False)
		self.parent.title(version)
		# Creating Menubar 
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
	
		#Generate UI
		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		self.init_UI()
		self.init_UI_Configuration()

	def Config_Init(self):
		self.Roaming = os.environ['APPDATA'] + '\\NX Automation'
		self.AppConfig = self.Roaming + '\\config.ini'
	
		if not os.path.isdir(self.Roaming):
			try:
				os.mkdir(self.Roaming)
			except OSError:
				print ("Creation of the directory %s failed" % self.Roaming)
		else:
			print('Roaming folder exist.')

	def Init_Folder(FolderPath):
		if not os.path.isdir(FolderPath):
			try:
				os.mkdir(FolderPath)
			except OSError:
				print ("Creation of the directory %s failed" % FolderPath)

	def Function_Get_TimeStamp():		
		now = datetime.now()
		timestamp = str(int(datetime.timestamp(now)))			
		return timestamp

	# UI init
	def init_UI(self):
	
		self.Generate_Automation_Execution_UI(self.DataCompare)
		#self.Generate_File_Comparision_UI(self.FileComparison)
		#self.Generate_Folder_Comparision_UI(self.FolderComparison)
		#self.Generate_Optimizer_UI(self.Optimizer)
		self.Generate_Debugger_UI(self.Process)
		
		# Debugger

	#UI Function

	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		file = Menu(menubar, tearoff = 0)
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		self.TAB_CONTROL = ttk.Notebook(self.parent)
		#Tab
		
		self.DataCompare = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.DataCompare, text= self.LanguagePack.Tab['StructuredCompare'])
		'''
		#Tab
		self.FileComparison = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.FileComparison, text= self.LanguagePack.Tab['FileComparison'])
		#Tab
		'''

		'''
		self.FolderComparison = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.FolderComparison, text= self.LanguagePack.Tab['FolderComparison'])
		#Tab
		self.Optimizer = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Optimizer, text= self.LanguagePack.Tab['Optimizer'])
		'''
		#Tab
		self.Process = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Process, text= self.LanguagePack.Tab['Debug'])
		

		self.TAB_CONTROL.pack(expand=1, fill="both")
		return

	def Generate_Automation_Execution_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_DB_Path = StringVar()
		self.Str_DB_Path.set('C:\\Users\\evan\\OneDrive - NEXON COMPANY\\[Demostration] V4 Gacha test\\DB\\db.xlsx')
		Label(Tab, text=  self.LanguagePack.Label['MainDB']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 130, state="readonly", textvariable=self.Str_DB_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=6, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_Test_Case_Path = StringVar()
		
		Label(Tab, text=  self.LanguagePack.Label['TestCaseList']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 130, state="readonly", textvariable=self.Str_Test_Case_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=6, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Test_Case_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
	
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['TestProject']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TestProject = AutocompleteCombobox(Tab)

		#self.TestProject = AutocompleteCombobox(Tab)
		self.TestProject.set_completion_list(['V4GB', "MSMG"])
		self.TestProject.Set_Entry_Width(30)
		self.TestProject.grid(row=Row, column=3, padx=5, pady=5, sticky=W)

		Row += 1
		Label(Tab, text=self.LanguagePack.Label['TestFeature']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TestFeature = AutocompleteCombobox(Tab)
		self.TestFeature.set_completion_list(['Gacha'])
		self.TestFeature.Set_Entry_Width(30)
		self.TestFeature.grid(row=Row, column=3, padx=5, pady=5, sticky=W)

		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Serial']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)

		self.TextSerial = Text(Tab, width=80, height=1, undo=True, wrap=WORD)
		self.TextSerial.insert("end", "RF8N12EQJBK")
		self.TextSerial.grid(row=Row, column=2, columnspan=4, padx=5, pady=5, sticky=W)

		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Execute_Script).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

	def Generate_Debugger_UI(self,Tab):
		Row = 1
		self.Debugger = Text(Tab, width=125, height=15, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)

	# Other function
	def Function_Import_DB(self, DB_Path, List_Sheet=[]):
		
		if DB_Path != None:
			if (os.path.isfile(DB_Path)):
				xlsx = load_workbook(DB_Path)
				MyDB = {}
				#Entry = {}

				for sheet in xlsx:
					sheetname = sheet.title.lower()
					if sheetname in List_Sheet:	
						DB_Name = sheetname
						
						Col_StringID = ""
						Col_String_EN = ""
						Col_String_KR = ""
						Col_Path = ""

						ws = xlsx[sheet.title]

						database = None

						for row in ws.iter_rows():
							for cell in row:
								if cell.value == "StringID":
									Col = cell.column_letter
									Row_ColID = cell.row
									Col_StringID = Col

								if cell.value == "String_EN":
									Col = cell.column_letter
									Col_String_EN = Col

								if cell.value == "String_KR":
									Col = cell.column_letter
									Col_String_KR = Col

								if cell.value == "Path":
									Col = cell.column_letter
									Col_Path = Col

									database = ws
									
								elif cell.value == "EN":
									EN_Coll = cell.column_letter
								if Col_StringID != "":
									database = ws
									break	
							if database!=  None:
								break		
	
						if database != None:
							for i in range(Row_ColID, database.max_row): 			
								StringID = database[Col_StringID + str(i+1)].value
								String_EN = database[Col_String_EN + str(i+1)].value
								String_KR = database[Col_String_KR + str(i+1)].value
								Path = database[Col_Path + str(i+1)].value
								Entry = DataEntry(String_EN, String_KR, Path)
								MyDB.StringID = Entry
				return MyDB
			else:
				return({})	
		else:
			return({})	

	def ImportTestCase(self, Test_Case_File_Path):
		print('Loading My Dictionary')
		if DatabasePath != None:
			if (os.path.isfile(DatabasePath)):
				xlsx = load_workbook(DatabasePath)
				DictList = []
				Dict = []
				for sheet in xlsx:
					sheetname = sheet.title.lower()
					if sheetname not in self.SpecialSheets:	
						EN_Coll = ""
						KR_Coll = ""
						database = None
						ws = xlsx[sheet.title]
						for row in ws.iter_rows():
							for cell in row:
								if cell.value == "KO":
									KR_Coll = cell.column_letter
									KR_Row = cell.row
									database = ws
								elif cell.value == "EN":
									EN_Coll = cell.column_letter
								if KR_Coll != "" and EN_Coll != "":
									DictList .append(sheet.title)
									break	
							if database!=  None:
								break		
						if database != None:
							for i in range(KR_Row, database.max_row): 
								KRAddress = KR_Coll + str(i+1)
								ENAddress = EN_Coll + str(i+1)
								#print('KRAddress', KRAddress)
								#print('ENAddress', ENAddress)
								KRCell = database[KRAddress]
								KRValue = KRCell.value
								ENCell = database[ENAddress]
								ENValue = ENCell.value
								if KRValue == None or ENValue == None or KRValue == 'KO' or ENValue == 'EN':
									continue
								elif KRValue != None and ENValue != None:
									Dict.append([KRValue, ENValue.lower()])
				print("Successfully load dictionary from: ", DictList)
				return Dict
			else:
				return([])	
		else:
			return([])	

	# Menu Function
	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Creator: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):

		self.Notice.set('Update app language...') 

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentToolkit'):
			config.add_section('DocumentToolkit')
			cfg = config['DocumentToolkit']	
		else:
			cfg = config['DocumentToolkit']

		cfg['applang']= language
		with open(self.AppConfig, 'w') as configfile:
			config.write(configfile)
		self.Notice.set('Config saved...')
		return

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()

	def Menu_Function_Open_Main_Guideline(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/Document+Toolkit")

	def Function_Correct_Path(self, path):
		return str(path).replace('/', '\\')
	
	def Function_Correct_EXT(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename, Obs_ext = os.path.splitext(baseName)
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def ErrorMsg(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def OpenOutput(self):
		Source = self.ListFile[0]

		Outputdir = os.path.dirname(Source)
		BasePath = str(os.path.abspath(Outputdir))
		subprocess.Popen('explorer ' + BasePath)
		
	def BtnLoadDocument(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("All type files","*.docx *.xlsx *.pptx *.msg"), ("Workbook files","*.xlsx *.xlsm"), ("Document files","*.docx"), ("Presentation files","*.pptx"), ("Email files","*.msg")), multiple = True)	
		if filename != "":
			self.ListFile = list(filename)
			self.CurrentSourceFile.set(str(self.ListFile[0]))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def BtnLoadRawSource(self):
		#filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files","*.xlsx *.xlsm *xlsb"), ("Document files","*.docx")), multiple = True)	
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.RawFile = FolderName
			self.RawSource.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
		'''
		if filename != "":
			self.RawFile = list(filename)
			self.RawSource.set(self.RawFile)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
		'''

	def BtnLoadTrackingSource(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("TM file","*.pkl"),),)	
		if filename != "":
			self.TrackingFile = self.CorrectPath(filename)
			print(self.TrackingFile)
			self.TrackingSource.set(self.CorrectPath(self.TrackingFile))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def BtnLoadRawTM(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("TM file","*.pkl"),), multiple = True)	
		if filename != "":
			self.RawTMFile = list(filename)
			Display = self.CorrectPath(self.RawTMFile[0])
			self.RawTMSource.set(Display)
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
	
	
	def BtnSelectColour(self):
		colorStr, self.BackgroundColor = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.BackgroundColor == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.BackgroundColor = 'ffff00'
		else:
			self.BackgroundColor = self.BackgroundColor.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_Browse_Old_Data_Folder(self):
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.OldDataTable = FolderName
			self.OldDataString.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_New_Data_Folder(self):
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.NewDataTable = FolderName
			
			self.NewDataString.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return	

	def BtnLoadDataLookupSource(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files *.xlsx"), ), multiple = True)	
		if filename != "":
			self.RawFile = list(filename)
			self.RawSource.set(self.RawFile)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def onExit(self):
		self.quit()


	def initGeneralSetting(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('DocumentToolkit'):
				cfg = config['DocumentToolkit']
			else:
				config['DocumentToolkit'] = {}
				cfg = config['DocumentToolkit']

			if config.has_option('DocumentToolkit', 'applang'):	
				self.AppLanguage = config['DocumentToolkit']['applang']
				print('Setting saved: ', self.AppLanguage)
			else:
				self.AppLanguage = 'en'
				print('Setting not saved: ', self.AppLanguage)

			#if config.has_option('Translator', 'Subscription'):
			#	self.Subscription = config['Translator']['Subscription']
			#else:
			#	self.Subscription = ''

		else:

			self.AppLanguage = 'en'

	def init_UI_Configuration(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('Document_Utility'):
				cfg = config['Document_Utility']
			else:
				config['Document_Utility'] = {}
				cfg = config['Document_Utility']

			if config.has_section('Comparision'):
				cfg = config['Comparision']
			else:
				config.add_section('Comparision')
				cfg = config['Comparision']

			
			'''
			if config.has_option('DocumentTranslator', 'turbo'):
				self.TurboTranslate.set(int(config['DocumentTranslator']['turbo']))
			else:
				self.TurboTranslate.set(0)
			'''
			
		else:
			self.Language = 'en'
			

		return	

	def SaveSetting(self):

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentTranslator'):
			config.add_section('DocumentTranslator')
			cfg = config['DocumentTranslator']	
		else:
			cfg = config['DocumentTranslator']

		cfg['lang']= str(self.Language.get())
		cfg['turbo']=  str(self.TurboTranslate.get())
		cfg['value']= str(self.DataOnly.get())
		cfg['filename']= str(self.TranslateFileName.get())
		cfg['sheetname']= str(self.TranslateSheetName.get())
		cfg['tmenable']= str(self.TMTranslate.get())
		cfg['tmupdate']= str(self.TMUpdate.get())
		cfg['sheetremoval']= str(self.SheetRemoval.get())
		#config['Document Translator'] = {'overwrite': self.OverWrite.get()}
		#config['Document Translator'] = {'skiprow': self.SkipEmptyRow.get()}
		#config['Document Translator'] = {'sheetremoval': self.SheetRemoval.get()}
		with open(appconfig, 'w') as configfile:
			config.write(self.AppConfig)
		self.Notice.set('Config saved...')
		return

	def GenerateTranslatorEngine(self):
		self.Notice.set(self.LanguagePack.ToolTips['AppInit'])
		if self.Language.get() == 1:
			to_language = 'ko'
			from_language = 'en'
		else:
			to_language = 'en'
			from_language = 'ko'

		self.p1 = Process(target=GenerateTranslator, args=(self.MyTranslator_Queue, self.TMManager, self.TrackingManager, self.MyTranslatorAgent, from_language, to_language, self.SubscriptionKey,))
		self.p1.start()
		self.after(DELAY1, self.GetMyTranslator)
		return

	def GenerateTranslatorEngineWithAlternativeDB(self):
		self.Notice.set(self.LanguagePack.ToolTips['AppInit'])
		if self.Language.get() == 1:
			to_language = 'ko'
			from_language = 'en'
		else:
			to_language = 'en'	
			from_language = 'ko'

		self.p1 = Process(target=GenerateTranslatorWithAlternativeDB, args=(self.MyTranslator_Queue, self.TMManager, self.TrackingManager, self.MyTranslatorAgent, from_language, to_language, self.SubscriptionKey, ))
		self.p1.start()
		self.after(DELAY1, self.GetMyTranslator)
		return
		


	def GetMyTranslator(self):
		try:
			self.MyTranslator = self.MyTranslator_Queue.get_nowait()
		except queue.Empty:
			self.after(DELAY1, self.GetMyTranslator)

		#print("self.MyTranslator: ", self.MyTranslator)	
		if self.MyTranslator != None:	
			self.TranslateBtn.configure(state=NORMAL)
			self.DictionaryStatus.set(str(len(self.MyTranslator.Dictionary)))
			self.ExceptionStatus.set(str(len(self.MyTranslator.Exception)))	
			self.TMStatus.set(str(len(self.MyTranslator.TranslationMemory)))



			self.DictionaryPath = self.MyTranslator.DictionaryPath
			self.AlternativeDictionaryPath = self.MyTranslator.AlternativeDictionaryPath
			self.ExceptionPath = self.MyTranslator.ExceptionPath
			self.TMPath = self.MyTranslator.TMPath
			
			self.CurrentAlternativeDB.set(self.AlternativeDictionaryPath)
			self.CurrentDB.set(self.DictionaryPath)
			self.Notice.set(self.LanguagePack.ToolTips['AppInitDone'])
			self.p1.join()
		else:
			self.Notice.set(self.LanguagePack.ToolTips['AppInit']) 
		return
	
	def TMTranslateModeToggle(self):
		if self.TMTranslate.get() == 1:
			self.MyTranslator.TMModeEnable(True)
		else:
			self.MyTranslator.TMModeEnable(False)

	def GetOptions(self):
		#Get and set language
		if self.Language.get() == 1:
			self.Options['to_language'] = 'ko'
			self.Options['LanguageName'] = self.LanguagePack.Option['Hangul']
			self.Options['from_language'] = 'en'
		else:
			self.Options['to_language'] = 'en'
			self.Options['LanguageName'] = self.LanguagePack.Option['English']
			self.Options['from_language'] = 'ko'
		
		self.MyTranslator.SetTargetLanguage(self.Options['to_language'])
		self.MyTranslator.SetSourceLanguage(self.Options['from_language'])
		self.Notice.set(self.LanguagePack.ToolTips['SetLanguage'] + self.Options['LanguageName'])
		#Set translator engine
		self.MyTranslator.SetTranslatorAgent(self.MyTranslatorAgent)

		
		#Add Subscription key
		self.MyTranslator.SetSubscriptionKey(self.SubscriptionKey)	

		#Set TM Update Mode
		if self.TMUpdate.get() == 1:
			self.MyTranslator.TMUpdateModeEnable(True)
		else:
			self.MyTranslator.TMUpdateModeEnable(False)

		#Set Predict mode 
		if self.TurboTranslate.get() == 1:
			self.MyTranslator.PredictModeEnable(True)
		else:
			self.MyTranslator.PredictModeEnable(False)

		#Set Data Mode
		if self.DataOnly.get() == 1:
			self.Options['DataMode']  = True
		else:
			self.Options['DataMode'] = False	

		#Set Skip Mode
		if self.SkipEmptyRow.get() == 1:
			self.Options['SkipMode'] = True
		else:
			self.Options['SkipMode'] = False

		#Set Sheet removal mode
		if self.SheetRemoval.get() == 1:
			self.Options['SheetRemovalMode'] = True
		else:
			self.Options['SheetRemovalMode'] = False

		# Set Update Mode
		if self.TMUpdate.get() == 1:
			self.Options['TMUpdateMode'] = True
		else:
			self.Options['TMUpdateMode'] = False



		#Set TM Translate Mode
		print('self.TMTranslate', self.TMTranslate.get())
		if self.TMTranslate.get() == 1:
			self.Options['MemoryTranslate']  = True
		else:
			self.Options['MemoryTranslate'] = False

		#Set Translate File name option
		if self.TranslateFileName.get() == 1:
			self.Options['TranslateFileName'] = True
		else:
			self.Options['TranslateFileName'] = False
		
		#Set Translate Sheet name option
		if self.TranslateSheetName.get() == 1:
			self.Options['TranslateSheetName'] = True
		else:
			self.Options['TranslateSheetName'] = False

		#Set Name fixer mode
		if self.FixCorruptFileName.get() == 1:
			self.Options['FixCorruptFileName'] = True
		else:
			self.Options['FixCorruptFileName'] = False	

		#Get sheet list
		try:
			Raw = self.SheetList.get("1.0", END).replace("\n", "").replace(" ", "")	
			self.Options['Sheet'] = Raw.split(",")
		except:
			self.Options['Sheet'] = []

		#Calculate Multiple value
		try:
			self.Options['Multiple'] = int(self.Multiple.get("1.0", END).replace("\n", "").replace(" ", ""))
		except:
			self.Options['Multiple'] = 8

		# Get Documents list
		self.Options['SourceDocument'] = self.ListFile
		if self.Options['SourceDocument'] == "":
			self.Notice.set(self.LanguagePack.ToolTips['SourceNotSelected']) 
			return
		else:
			self.Notice.set(self.LanguagePack.ToolTips['DocumentLoad']) 


	def BtnExportTrackingData(self):
		SourceDocument = self.TrackingFile

		self.processTracking = Process(target=ExportData, args=(SourceDocument, self.StatusQueue,))
		self.processTracking.start()
		self.after(DELAY1, self.GetTrackingStatus)	

	def GetTrackingStatus(self):
		if (self.processTracking.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetTrackingStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.processTracking.terminate()

	def BtnBrowseTMSource(self):
		filename = filedialog.askopenfilename(title = "Select TM file",filetypes = (("Translation Memory","*.pkl"), ), multiple = True)	
		if filename != "":
			self.TMSourceList = list(filename)
			for Path in self.TMSourceList:
				Path =  self.CorrectPath(Path)
			
			Display = str(self.CorrectPath(self.TMSourceList[0]))

			self.TMSource.set(Display)
		else:
			self.Notice.set("No document is selected")
			return

	def BtnMergeTM(self):
		if self.TMSource == None:
			self.Notice.set("Please select TM file before starting")
		#elif not os.path.isfile(self.DictionarySource):
		#	self.Notice.set("TM files has been removed")
		else:	
			filename = filedialog.asksaveasfilename(title = "Save file to", filetypes = (("Translation Memory", "*.pkl"),),)
			newPath = self.CorrectExt(filename, "pkl")
			print("Save TM to: ", newPath)
			'''
			Outputdir = os.path.dirname(filename)
			baseName = os.path.basename(filename)
			sourcename, ext = os.path.splitext(baseName)

			newPath = Outputdir + '/'+ sourcename + '.pkl'
			'''
			if filename == "":
				self.Notice.set("Please enter a file name.")
				self.ErrorMsg("Please enter a file name.")
			else:
				
				SourceDocument = self.TMSourceList

				self.p4 = Process(target=MergeTMSource, args=(SourceDocument, newPath, self.StatusQueue,))
				self.p4.start()
				self.after(DELAY1, self.GetMergeTMStatus)	

	def GetMergeTMStatus(self):
		if (self.p4.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

	def OpenOptimizerFolder(self):
		try:
			SourceDocument = self.RawFile
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return	
		BasePath = str(os.path.abspath(self.RawFile))
		subprocess.Popen('explorer ' + BasePath)
	
	def BtnOptimizeXLSX(self):
		try:
			SourceDocument = self.RawFile
		except AttributeError:
			self.ErrorMsg('Please select source folder.')

		try:
			while True:
				percent = self.Process_Queue.get_nowait()
				#print("Remain percent: ", percent)
		except queue.Empty:
			pass
		self.Optimize_Process = Process(target=Function_Optimize_XLSX, args=(SourceDocument, self.Status_Queue,self.Process_Queue,))
		self.Optimize_Process.start()
		self.after(DELAY1, self.GetOptimizeStatus)	

	def GetOptimizeStatus(self):
		if (self.Optimize_Process.is_alive()):
			try:
				percent = self.Process_Queue.get(0)
				self.Optimize_Progressbar["value"] = percent
				self.Optimize_Progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass
			
			
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Optimize_Process.terminate()

	def BtnOptimizeTM(self):
		SourceDocument = self.RawTMFile

		self.p4 = Process(target=OptimizeTM, args=(SourceDocument, self.Status_Queue,))
		self.p4.start()
		self.after(DELAY1, self.GetOptimizeTMStatus)	

	def GetOptimizeTMStatus(self):
		if (self.p4.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

	def BtnCompareLookupData(self):
		#OldDocument = self.OldDataTable
		LookupData = self.DataLookUpSource
		if self.AllData.get() == 0:
			DataOnly = True
		else:
			DataOnly = False

		self.p4 = Process(target=CompareTableData, args=(self.StatusQueue, LookupData, LookupValue, Mode, DataOnly,))
		self.p4.start()
		self.after(DELAY1, self.GetCompareStatus)

###########################################################################################
	def BtnCompareDataTable(self):

		try:
			OldDocument = self.OldDataTable
			NewDocument = self.NewDataTable
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return

		try:
			self.BackgroundColor
		except:
			self.BackgroundColor = 'ffff00'	
		if self.BackgroundColor == False or self.BackgroundColor == None:
			self.BackgroundColor = 'ffff00'
		print('self.BackgroundColor: ', self.BackgroundColor)

		try:
			self.BackgroundColor
		except:
			self.BackgroundColor = 'ffff00'	
		if self.BackgroundColor == False or self.BackgroundColor == None:
			self.BackgroundColor = 'ffff00'
		print('self.BackgroundColor: ', self.BackgroundColor)

		if self.AllData.get() == 0:
			DataOnly = True
		else:
			DataOnly = False

		self.p4 = Process(target=Compare_Folder_Data, args=(self.Status_Queue, self.Process_Queue, OldDocument, NewDocument, self.BackgroundColor, DataOnly,))
		self.p4.start()
		self.after(DELAY1, self.GetCompareStatus)	

	def GetCompareStatus(self):
		if (self.p4.is_alive()):
			try:
				percent = self.Process_Queue.get(0)
				self.CompareProgressbar["value"] = percent
				self.CompareProgressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetCompareStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

###########################################################################################

###########################################################################################
	def Btn_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Background_Color == None:
			self.Error('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	def Btn_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.Error('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_Browse_DB_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)	
		if filename != "":
			self.DB_Path = self.Function_Correct_Path(filename)
			self.Str_DB_Path.set(self.DB_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_Test_Case_File(self):
		
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)
		
		if filename != "":
			self.Test_Case_Path = self.Function_Correct_Path(filename)
			self.Str_Test_Case_Path.set(self.Test_Case_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Execute_Script(self):
		DB = self.Str_DB_Path.get()
		
		Test_Case = self.Test_Case_Path
		Serial = self.TextSerial.get("1.0", END).replace('\n','')
		#MyDB = self.Function_Import_DB(self.DB_Path)

		
		self.Automation_Processor = Process(target=Function_Execute_Script, args=(self.Status_Queue, self.Process_Queue, Serial, DB, Test_Case,))
		#self.Data_Compare_Process = Process(target=Old_Function_Compare_Excel, args=(self.Status_Queue, self.Process_Queue, Old_File, New_File, Output_Result, Sheet_Name, Index_Col, self.Background_Color, self.Font_Color,))
		self.Automation_Processor.start()
		self.after(DELAY1, self.Wait_For_Automation_Processor)	

	def Wait_For_Automation_Processor(self):
		if (self.Automation_Processor.is_alive()):
			'''
			try:
				percent = self.ProcessQueue.get(0)
				self.CompareProgressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			'''
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_Automation_Processor)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Automation_Processor.terminate()

###########################################################################################

###########################################################################################

def Function_Execute_Script(
		Status_Queue, Process_Queue, Serial_Nummber, DB_Path, Test_Case_Path, **kwargs
):

	Status_Queue.put("Importing DB")
	#DB = Function_Import_DB(DB_Path)
	
	#Serial = "RF8N12EQJBK"

	Test = V4Test(Serial_Nummber, DB_Path)

	#Test.Nav_V4Shop()

	Test.Tap_Item('UI_V4Shop')
	Test.Tap_Item('UI_BurgerMenu')
	Test.Tap_Item('UI_Inventory')
	Test.Tap_Item('UI_V4Shop')
	Test.Tap_Item('UI_Exit')
	Test.Tap_Item('UI_BurgerMenu')


###########################################################################################

def main():
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue

	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager

	Glossary_Manager(root, Queue = My_Queue, Manager = My_Manager,)
	root.mainloop()  


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
