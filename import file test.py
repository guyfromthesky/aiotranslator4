import pkgutil
from openpyxl import load_workbook
import io

help_bin = pkgutil.get_data( 'Document Translator', 'scr/version.txt' )
help_utf = help_bin.decode('UTF-8', 'ignore')
print(help_utf)

language_pack = pkgutil.get_data( 'Document Translator', 'scr/language.xlsx' )
xlsx = io.BytesIO(language_pack)

wb = load_workbook(xlsx)
for sheet in wb:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value != None:
                current_string = str(cell.value)
                print('current_string', current_string)