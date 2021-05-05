#System variable and io handling
import sys, getopt
import os
import multiprocessing
from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp
import time
from datetime import datetime
import configparser

#GUI
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkinter import colorchooser
from tkinter import simpledialog
#from tkinter import style

from openpyxl import load_workbook, worksheet, Workbook
import csv

from libs.aiotranslator import Translator
from libs.aioconfigmanager import ConfigLoader
from google.cloud import logging
#MyTranslatorAgent = 'google'
Tool = "DB converter"

from libs.version import get_version
rev = 1300
version = Tool  + " " +  get_version(rev)
DELAY1 = 20

#import testscript
class AutocompleteCombobox(Combobox):

	def set_completion_list(self, completion_list):
		"""Use our completion list as our drop down selection menu, arrows move through menu."""
		self._completion_list = sorted(completion_list, key=str.lower) # Work with a sorted list
		self._hits = []
		self._hit_index = 0
		self.position = 0
		self.bind('<KeyRelease>', self.handle_keyrelease)
		self['values'] = self._completion_list  # Setup our popup menu
		#self._w = 10
		self.delete(0,END)	

	def Set_Entry_Width(self, width):
		self.configure(width=width)

	def Set_DropDown_Width(self, width):
		print('Change size: ', width)
		style = Style()
		style.configure('TCombobox', postoffset=(0,0,width,0))
		self.configure(style='TCombobox')

	def autocomplete(self, delta=0):
		"""autocomplete the Combobox, delta may be 0/1/-1 to cycle through possible hits"""
		if delta: # need to delete selection otherwise we would fix the current position
			self.delete(self.position, END)
		else: # set position to end so selection starts where textentry ended
			self.position = len(self.get())
		# collect hits
		_hits = []
		for element in self._completion_list:
			if element.lower().startswith(self.get().lower()): # Match case insensitively
				_hits.append(element)
		# if we have a new hit list, keep this in mind
		if _hits != self._hits:
			self._hit_index = 0
			self._hits=_hits
		# only allow cycling if we are in a known hit list
		if _hits == self._hits and self._hits:
			self._hit_index = (self._hit_index + delta) % len(self._hits)
		# now finally perform the auto completion
		if self._hits:
			self.delete(0,END)
			self.insert(0,self._hits[self._hit_index])
			self.select_range(self.position,END)

	def handle_keyrelease(self, event):
		"""event handler for the keyrelease event on this widget"""
		if event.keysym == "BackSpace":
			self.delete(self.index(INSERT), END)
			self.position = self.index(END)
		if event.keysym == "Left":
			if self.position < self.index(END): # delete the selection
				self.delete(self.position, END)
			else:
				self.position = self.position-1 # delete one character
				self.delete(self.position, END)
		if event.keysym == "Right":
			self.position = self.index(END) # go to end (no selection)
		if len(event.keysym) == 1:
			self.autocomplete()

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class Automation_Execuser(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		#super().__init__()
		self.parent = Root 

		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']

		self.Manager = Manager['Default_Manager']

		#self.Config_Init()

		self.Options = {}

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		self.ButtonWidth = 20
		self.HalfButtonWidth = 15

		self.DB_Path = ""
		self.Test_Case_Path = ""
		
		self.App_LanguagePack = {}

		self.init_App_Setting()

		if self.AppLanguage != 'kr':
			from libs.languagepack import LanguagePackEN as LanguagePack
		else:
			from libs.languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack

		# Init function

		self.parent.resizable(False, False)
		self.parent.title(version)
		# Creating Menubar 
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
	
		#Generate UI
		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		self.init_UI()

		self.Init_Gloss_List()

	def Init_Gloss_List(self):
		MyTranslator = Translator('en', 'ko', False, False, False, None,)
		self.Text_GlossaryID.set_completion_list(MyTranslator.GlossaryList)

	def init_App_Setting(self):

		self.LicensePath = StringVar()
		self.DictionaryPath = StringVar()
		self.TMPath = StringVar()

		self.CurrentDataSource = StringVar()
		self.Notice = StringVar()
		self.DictionaryStatus = StringVar()
		
		self.TMStatus  = StringVar()
		self.HeaderStatus = StringVar()

		self.AppConfig = ConfigLoader()
		self.Configuration = self.AppConfig.Config
		self.AppLanguage  = self.Configuration['Bug_Writer']['app_lang']
		self.SkipTestInfo = self.Configuration['Bug_Writer']['test_info_inable']
		self.UseSimpleTemplate = self.Configuration['Bug_Writer']['use_simple_template']
		
		self.LicensePath.set(self.Configuration['License_File']['path'])
		os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = self.Configuration['License_File']['path']
		self.DictionaryPath.set(self.Configuration['Database']['path'])
		self.TMPath.set(self.Configuration['TranslationMemory']['path'])

		self.GlossaryID = self.Configuration['Glossary_ID']['value']


	def Init_Folder(self, FolderPath):
		if not os.path.isdir(FolderPath):
			try:
				os.mkdir(FolderPath)
			except OSError:
				print ("Creation of the directory %s failed" % FolderPath)

	def Function_Get_TimeStamp(self):		
		now = datetime.now()
		timestamp = str(int(datetime.timestamp(now)))			
		return timestamp

	# UI init
	def init_UI(self):
	
		#self.Generate_Automation_Execution_UI(self.DataCompare)
		self.Generate_Glossary_Manager_UI(self.GlossaryManager)
		self.Generate_TranslateSetting_UI(self.TranslateSetting)
		#self.Generate_Glossary_Create_UI(self.GlossaryCreator)
		#self.Generate_File_Comparision_UI(self.FileComparison)
		#self.Generate_Folder_Comparision_UI(self.FolderComparison)
		#self.Generate_Optimizer_UI(self.Optimizer)
		#self.Generate_Debugger_UI(self.Process)
		
		# Debugger

	#UI Function

	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		file = Menu(menubar, tearoff = 0)
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		self.TAB_CONTROL = ttk.Notebook(self.parent)
		#Tab

		
		self.GlossaryManager = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.GlossaryManager, text= 'Glossary Uploader')	

		self.TranslateSetting = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.TranslateSetting, text= 'Translate Setting')	

		#self.GlossaryCreator = ttk.Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.GlossaryCreator, text= 'Glossary Creator')	


		#self.DataCompare = ttk.Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.DataCompare, text= 'DB converter')


		#Tab
		#self.Process = ttk.Frame(self.TAB_CONTROL)
		#self.TAB_CONTROL.add(self.Process, text= self.LanguagePack.Tab['Debug'])
		

		self.TAB_CONTROL.pack(expand=1, fill="both")
		return

	
	def Generate_Glossary_Manager_UI(self, Tab):
		
		
		Row =1
		self.Str_DB_Path = StringVar()
		#self.Str_DB_Path.set('C:\\Users\\evan\\OneDrive - NEXON COMPANY\\[Demostration] V4 Gacha test\\DB\\db.xlsx')
		Label(Tab, text=  self.LanguagePack.Label['MainDB']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 130, state="readonly", textvariable=self.Str_DB_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=6, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['ProjectKey']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_GlossaryID = AutocompleteCombobox(Tab)
		self.Text_GlossaryID.Set_Entry_Width(30)
		self.Text_GlossaryID.set_completion_list([])
		self.Text_GlossaryID.grid(row=Row, column=3, columnspan=2, padx=5, pady=5, stick=W)

		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Execute_Script).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=E)

		
		
		Row += 1
		self.Debugger = Text(Tab, width=125, height=10, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)

	def Generate_TranslateSetting_UI(self, Tab):
		Row = 1
		Label(Tab, text= self.LanguagePack.Label['LicensePath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TextLicensePath = Entry(Tab,width = 120, state="readonly", textvariable=self.LicensePath)
		self.TextLicensePath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.HalfButtonWidth, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_License_Path).grid(row=Row, column=8, columnspan=2, padx=5, pady=5, sticky=E)

		'''
		Row += 1
		Label(Tab, width = 30, text= self.LanguagePack.Label['DBSourcePath']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		self.TextFilePath = Entry(Tab, width = 120, text="Select your document", state="readonly", textvariable=self.DictionaryPath.get())
		self.TextFilePath.grid(row=Row, column=3, columnspan=7, padx=5, pady=5, sticky=W)
		#Button(Tab, width = 20, text=  self.LanguagePack.Button['Browse'], command= self.BtnLoadDocument).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.HalfButtonWidth, text= self.LanguagePack.Button['Save'], command= self.SaveProjectKey).grid(row=Row, column=8, columnspan=2, padx=5, pady=5, sticky=E)
		'''

	def Generate_Glossary_Create_UI(self, Tab):
		
		Row =1
		self.Str_DB_Path = StringVar()
		#self.Str_DB_Path.set('C:\\Users\\evan\\OneDrive - NEXON COMPANY\\[Demostration] V4 Gacha test\\DB\\db.xlsx')
		Label(Tab, text=  self.LanguagePack.Label['MainDB']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 130, state="readonly", textvariable=self.Str_DB_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=6, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)

		Row += 1
		Label(Tab, text= self.LanguagePack.Label['ProjectKey']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_New_GlossaryID = Text(Tab, width=50, height=1, undo=True, wrap=WORD, )
		self.Text_New_GlossaryID.grid(row=Row, column=3, columnspan=2, padx=5, pady=5, stick=W)

		Row += 1
		Label(Tab, text= "URI").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_URI_ID = Text(Tab, width=50, height=1, undo=True, wrap=WORD, )
		self.Text_URI_ID.grid(row=Row, column=3, columnspan=2, padx=5, pady=5, stick=W)

		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Execute_Creator_Script).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=E)
		Row += 1

		self.Debugger = Text(Tab, width=125, height=10, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)


	def Generate_Debugger_UI(self,Tab):
		Row = 1
		self.Debugger = Text(Tab, width=125, height=15, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)

	# Menu Function
	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Creator: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):

		self.Notice.set('Update app language...') 

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentToolkit'):
			config.add_section('DocumentToolkit')
			cfg = config['DocumentToolkit']	
		else:
			cfg = config['DocumentToolkit']

		cfg['applang']= language
		with open(self.AppConfig, 'w') as configfile:
			config.write(configfile)
		self.Notice.set('Config saved...')
		return

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()

	def Menu_Function_Open_Main_Guideline(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/Document+Toolkit")

	def Function_Correct_Path(self, path):
		return str(path).replace('/', '\\')
	
	def Function_Correct_EXT(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename, Obs_ext = os.path.splitext(baseName)
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def Btn_Select_License_Path(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectDB'],filetypes = (("JSON files","*.json" ), ), )	
		if filename != "":
			LicensePath = self.CorrectPath(filename)
			self.AppConfig.Save_Config(self.AppConfig.Translator_Config_Path, 'License_File', 'path', LicensePath, True)
			os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = LicensePath
			self.LicensePath.set(LicensePath)
		else:
			self.Notice.set("No file is selected")

	def ErrorMsg(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def Confirm_Popup(self, Request, Message):
		MsgBox = simpledialog.askstring(title="Input project ID", prompt="What's your Project ID?")

		if MsgBox == Request:
			return True
		else:
			return False

	def onExit(self):
		self.quit()

	

	def SaveSetting(self):

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentTranslator'):
			config.add_section('DocumentTranslator')
			cfg = config['DocumentTranslator']	
		else:
			cfg = config['DocumentTranslator']

		cfg['lang']= str(self.Language.get())
		cfg['turbo']=  str(self.TurboTranslate.get())
		cfg['value']= str(self.DataOnly.get())
		cfg['filename']= str(self.TranslateFileName.get())
		cfg['sheetname']= str(self.TranslateSheetName.get())
		cfg['tmenable']= str(self.TMTranslate.get())
		cfg['tmupdate']= str(self.TMUpdate.get())
		cfg['sheetremoval']= str(self.SheetRemoval.get())
		#config['Document Translator'] = {'overwrite': self.OverWrite.get()}
		#config['Document Translator'] = {'skiprow': self.SkipEmptyRow.get()}
		#config['Document Translator'] = {'sheetremoval': self.SheetRemoval.get()}
		with open(appconfig, 'w') as configfile:
			config.write(self.AppConfig)
		self.Notice.set('Config saved...')
		return


###########################################################################################

###########################################################################################

	def Btn_Browse_DB_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)	
		if filename != "":
			self.DB_Path = self.Function_Correct_Path(filename)
			self.Str_DB_Path.set(self.DB_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_Glossary_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.svg"), ), multiple = False)	
		if filename != "":
			self.DB_Path = self.Function_Correct_Path(filename)
			self.Str_DB_Path.set(self.DB_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Execute_Creator_Script(self):
		DB = self.Str_DB_Path.get()
		Glossary_ID = self.Text_New_GlossaryID.get()
		URI = self.Text_URI_ID.get()

		self.Automation_Processor = Process(target=Function_Execute_Create_Script, args=(self.Status_Queue, DB, Glossary_ID, URI,))
		self.Automation_Processor.start()
		self.after(DELAY1, self.Wait_For_Automation_Creator_Processor)	

	def Wait_For_Automation_Creator_Processor(self):
		if (self.Automation_Processor.is_alive()):
			'''
			try:
				percent = self.ProcessQueue.get(0)
				self.CompareProgressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			'''
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_Automation_Creator_Processor)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Wait_For_Automation_Creator_Processor.terminate()

	def Btn_Execute_Script(self):
		DB = self.Str_DB_Path.get()
		Glossary_ID = self.Text_GlossaryID.get()
		
		result = self.Confirm_Popup(Glossary_ID, 'Are you sure you want to replace the DB of '+ Glossary_ID + "?")
		
		if result == True:
			
			self.Automation_Processor = Process(target=Function_Execute_Script, args=(self.Status_Queue, DB, Glossary_ID,))
			self.Automation_Processor.start()
			self.after(DELAY1, self.Wait_For_Automation_Processor)	

	def Wait_For_Automation_Processor(self):
		if (self.Automation_Processor.is_alive()):
			'''
			try:
				percent = self.ProcessQueue.get(0)
				self.CompareProgressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			'''
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_Automation_Processor)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Automation_Processor.terminate()

###########################################################################################

###########################################################################################
def Function_Execute_Script(Status_Queue, DB_Path, Glossary_ID, **kwargs):

	Output = Function_Create_CSV_DB(Status_Queue, DB_Path)
	Status_Queue.put("CSV DB created:" + str(Output))
	if Glossary_ID != '':
		
		client = logging.Client()
		log_name = 'db-update'
		logger = client.logger(log_name)
		try:
			account = os.getlogin()
		except:
			account = 'Anonymous'
		
		try:
			name = os.environ['COMPUTERNAME']
		except:
			name = 'Anonymous'	

		text_log = account + ', ' + name + ', ' + Glossary_ID + ', ' + DB_Path
		logger.log_text(text_log)

		myTranslator = Translator('ko', 'en', None, None, False, False, False, Glossary_ID, )
		
		myTranslator.Update_Glob(Glossary_ID, Output)

		Status_Queue.put("DB updated.")

def Function_Execute_Create_Script(Status_Queue, DB_Path, Glossary_ID, URI, **kwargs):

	Output = Function_Create_CSV_DB(Status_Queue, DB_Path)
	Status_Queue.put("CSV DB created:" + str(Output))
	myTranslator = Translator('ko', 'en', None, None, False, False, False, Glossary_ID, )
	
	myTranslator.Update_Glob(Glossary_ID, Output)	
	Status_Queue.put("DB created.")

def Function_Update_Glossary(self):

	return

def Function_Create_CSV_DB(
		Status_Queue, DB_Path, **kwargs
):

	DatabasePath = DB_Path

	Outputdir = os.path.dirname(DatabasePath)
	baseName = os.path.basename(DatabasePath)
	sourcename, ext = os.path.splitext(baseName)

	#output_file = Outputdir + '/' + sourcename + '_SingleFile.xlsx'
	output_file_csv = Outputdir + '/' + sourcename + '.csv'
	SpecialSheets = ['info']

	RowCount = 0

	if DatabasePath != None:
		if (os.path.isfile(DatabasePath)):
			xlsx = load_workbook(DatabasePath)
			DictList = []
			Dict = []
			with open(output_file_csv, 'w', newline='', encoding='utf_8_sig') as csv_file:
				writer = csv.writer(csv_file, delimiter=',')
				writer.writerow(['Description', 'ko', 'en'])
				for sheet in xlsx:
					sheetname = sheet.title.lower()
					
					if sheetname not in SpecialSheets:	
					
						EN_Coll = ""
						KR_Coll = ""
						database = None
						ws = xlsx[sheet.title]
						for row in ws.iter_rows():
							for cell in row:
								if cell.value == "KO":
									KR_Coll = cell.column_letter
									KR_Row = cell.row
									database = ws
								elif cell.value == "EN":
									EN_Coll = cell.column_letter
								if KR_Coll != "" and EN_Coll != "":
									DictList.append(sheet.title)
									break	
							if database!=  None:
								break		

						if database != None:
							
							for i in range(KR_Row, database.max_row): 
								KRAddress = KR_Coll + str(i+1)
								ENAddress = EN_Coll + str(i+1)
								KRCell = database[KRAddress]
								KRValue = KRCell.value
								ENCell = database[ENAddress]
								ENValue = ENCell.value
								if KRValue in [None, 'KO'] or ENValue in [None, 'EN']:
									continue
								elif KRValue not in  ["", 'KO'] and ENValue not in ["", 'EN']:
									KRValue = KRCell.value.replace('\r', '').replace('\n', '')
									ENValue = ENCell.value.replace('\r', '').replace('\n', '')
									if sheetname != 'header':
										ENValue = ENValue.lower()
									writer.writerow([sheetname, KRValue, ENValue])
									RowCount+=1
					elif sheetname == 'info':
						ws = xlsx[sheet.title]
						for row in ws.iter_rows():
							for cell in row:
								if cell.value == "i_version":
									temp_Col = chr(ord(cell.column_letter) + 1)
									temp_Row = cell.row
									temp_Add = temp_Col + str(temp_Row)	
									temp_Cel = ws[temp_Add]
									temp_Val = temp_Cel.value
									writer.writerow(['info', 'i_version', temp_Val])
								elif cell.value == "i_date":
									temp_Col = chr(ord(cell.column_letter) + 1)
									temp_Row = cell.row
									temp_Add = temp_Col + str(temp_Row)	
									temp_Cel = ws[temp_Add]
									temp_Val = temp_Cel.value
									writer.writerow(['info', 'i_date', temp_Val])

		Status_Queue.put("Successfully load dictionary from: " +  str(DictList))
	#Status_Queue.put("Outputfile: " +  str(output_file_csv))
	#Status_Queue.put("Total pair: " +  str(RowCount))
	return output_file_csv


###########################################################################################

def main():
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue

	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager

	Automation_Execuser(root, Queue = My_Queue, Manager = My_Manager,)
	root.mainloop()  


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
